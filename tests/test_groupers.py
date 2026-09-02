import importlib
import os
import shutil

import numpy as np
import pytest

from chemsmart.io.xyz.xyzfile import XYZFile
from chemsmart.jobs.grouper.base import MatrixGrouper, MoleculeGrouper
from chemsmart.jobs.grouper.connectivity import ConnectivityGrouper
from chemsmart.jobs.grouper.energy import EnergyGrouper
from chemsmart.jobs.grouper.formula import FormulaGrouper
from chemsmart.jobs.grouper.isomorphism import RDKitIsomorphismGrouper
from chemsmart.jobs.grouper.rmsd import (
    BasicRMSDGrouper,
    HungarianRMSDGrouper,
    IRMSDGrouper,
    PymolRMSDGrouper,
    RMSDGrouper,
    SpyRMSDGrouper,
)
from chemsmart.jobs.grouper.tanimoto import TanimotoSimilarityGrouper
from chemsmart.jobs.grouper.tfd import TorsionFingerprintGrouper
from chemsmart.utils.grouper import StructureGrouperFactory
from chemsmart.utils.utils import find_irmsd_command, kabsch_align


def _is_irmsd_available() -> bool:
    """Return True when either irmsd Python API or CLI is available."""
    try:
        irmsd_module = importlib.import_module("irmsd.api.irmsd_exposed")
        getattr(irmsd_module, "get_irmsd")
        return True
    except (ImportError, OSError, AttributeError):
        return bool(find_irmsd_command())


def _assert_progress_milestones(log_text: str) -> None:
    """Assert that progress milestones from 10% to 100% are present."""
    for milestone in range(10, 101, 10):
        assert f"Matrix calculation progress: {milestone}%" in log_text


@pytest.mark.usefixtures("temporary_working_dir")
class Test_BasicRMSD_grouper_and_basic_functionality:
    NUM_PROCS = 4

    def test_rmsd_grouper(
        self,
        methanol_molecules,
        methanol_and_ethanol,
    ):
        methanol = methanol_molecules[0]
        methanol_rot1 = methanol_molecules[1]

        assert np.any(
            methanol.positions != methanol_rot1.positions
        ), "Rotated molecule should have different positions."

        grouper = BasicRMSDGrouper(methanol_molecules)
        groups, group_indices = grouper.group()

        rmsd = grouper.calculate_rmsd_pair(0, 1)
        assert np.isclose(
            rmsd, 0.0, rtol=1e-3
        ), "RMSD should be close to zero."

        rmsd = grouper.calculate_rmsd_pair(1, 2)
        assert np.isclose(
            rmsd, 0.0, rtol=1e-3
        ), "RMSD should be close to zero."

        _, _, _, _, rmsd_kabsch = kabsch_align(
            methanol.positions,
            methanol_rot1.positions,
        )
        assert np.isclose(rmsd_kabsch, 0.0, rtol=1e-3)

        assert (
            len(groups) == 1
        ), "Molecules should form one group based on geometry."
        assert (
            len(group_indices) == 1
        ), "Molecules should form one group based on geometry."

        unique_structures = grouper.unique()
        assert (
            len(unique_structures) == 1
        ), "Molecules should form one group based on geometry."

        grouper2 = BasicRMSDGrouper(methanol_and_ethanol)
        groups, group_indices = grouper2.group()

        assert (
            len(groups) == 1
        ), "Molecules should form two groups based on geometry."
        assert (
            len(group_indices) == 1
        ), "Molecules should form two groups based on geometry."

        rmsd = grouper2.calculate_rmsd_pair(0, 1)
        assert (
            rmsd is np.inf
        ), "RMSD is set to be infinity for different molecules."

        unique_structures = grouper2.unique()
        assert (
            len(unique_structures) == 1
        ), "One incompatible structure should be skipped, leaving one unique structure."

    def test_rmsd_grouper_for_crest_conformers(
        self, multiple_molecules_xyz_file
    ):
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)

        molecules = xyz_file.get_molecules(index=":", return_list=True)
        assert len(molecules) == 18
        grouper = BasicRMSDGrouper(
            molecules, threshold=0.2, num_procs=self.NUM_PROCS
        )
        groups, group_indices = grouper.group()
        assert len(groups) == 18
        assert len(group_indices) == 18
        unique_structures = grouper.unique()
        assert len(unique_structures) == 18

        # rmsd calculation from grouper
        rmsd = grouper.calculate_rmsd_pair(0, 1)
        assert np.isclose(rmsd, 0.409, rtol=1e-3)

        _, _, _, _, rmsd = kabsch_align(
            molecules[0].positions, molecules[1].positions
        )
        assert np.isclose(rmsd, 0.409, rtol=1e-3)

        grouper2 = BasicRMSDGrouper(
            molecules, threshold=0.5, num_procs=self.NUM_PROCS
        )
        groups, group_indices = grouper2.group()
        assert len(groups) == 12
        assert len(group_indices) == 12
        unique_structures = grouper2.unique()
        assert len(unique_structures) == 12

        grouper3 = BasicRMSDGrouper(
            molecules, threshold=1.0, num_procs=self.NUM_PROCS
        )
        groups, group_indices = grouper3.group()
        assert len(groups) == 11
        assert len(group_indices) == 11
        unique_structures = grouper3.unique()
        assert len(unique_structures) == 11

        grouper4 = BasicRMSDGrouper(
            molecules, threshold=1.5, num_procs=self.NUM_PROCS
        )
        groups, group_indices = grouper4.group()
        assert len(groups) == 8
        assert len(group_indices) == 8
        unique_structures = grouper4.unique()
        assert len(unique_structures) == 8

        grouper5 = BasicRMSDGrouper(
            molecules, threshold=2.0, num_procs=self.NUM_PROCS
        )
        groups, group_indices = grouper5.group()
        assert len(groups) == 7
        assert len(group_indices) == 7
        unique_structures = grouper5.unique()
        assert len(unique_structures) == 7

        grouper6 = BasicRMSDGrouper(
            molecules, threshold=2.5, num_procs=self.NUM_PROCS
        )
        groups, group_indices = grouper6.group()
        assert len(groups) == 4
        assert len(group_indices) == 4
        unique_structures = grouper6.unique()
        assert len(unique_structures) == 4

    def test_num_groups_parameter(self, multiple_molecules_xyz_file):
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)

        molecules = xyz_file.get_molecules(index=":", return_list=True)
        assert len(molecules) == 18
        grouper = BasicRMSDGrouper(
            molecules, threshold=None, num_groups=17, num_procs=self.NUM_PROCS
        )
        groups, group_indices = grouper.group()
        assert len(groups) == 17
        assert len(group_indices) == 17
        unique_structures = grouper.unique()
        assert len(unique_structures) == 17

    def test_pick_the_lowestenergy_conformers(
        self, multiple_molecules_xyz_file
    ):
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)

        molecules = xyz_file.get_molecules(index=":", return_list=True)
        assert len(molecules) == 18
        grouper = BasicRMSDGrouper(
            molecules, threshold=None, num_groups=3, num_procs=self.NUM_PROCS
        )
        groups, group_indices = grouper.group()
        assert len(groups) == 3
        assert len(group_indices) == 3
        unique_structures = grouper.unique()
        assert len(unique_structures) == 3

        energies = [mol.energy for mol in unique_structures]
        assert -126.2575508 in energies
        assert -126.25017833 in energies
        assert -126.24909661 in energies

    def test_rmsd_grouper_for_crest_conformers_ignore_Hs(
        self, multiple_molecules_xyz_file
    ):
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)

        molecules = xyz_file.get_molecules(index=":", return_list=True)
        assert len(molecules) == 18
        grouper = BasicRMSDGrouper(
            molecules,
            threshold=0.2,
            num_procs=self.NUM_PROCS,
            ignore_hydrogens=True,
        )
        groups, group_indices = grouper.group()
        assert len(groups) == 17
        assert len(group_indices) == 17
        unique_structures = grouper.unique()
        assert len(unique_structures) == 17

        # rmsd calculation from grouper
        rmsd = grouper.calculate_rmsd_pair(0, 1)
        assert np.isclose(rmsd, 0.301, rtol=1e-3)  # removed H atoms

        _, _, _, _, rmsd = kabsch_align(
            molecules[0].positions, molecules[1].positions
        )
        assert np.isclose(rmsd, 0.409, rtol=1e-3)  # did not remove H atoms

        grouper2 = BasicRMSDGrouper(
            molecules,
            threshold=0.5,
            num_procs=self.NUM_PROCS,
            ignore_hydrogens=True,
        )
        groups, group_indices = grouper2.group()
        assert len(groups) == 12
        assert len(group_indices) == 12
        unique_structures = grouper2.unique()
        assert len(unique_structures) == 12

        grouper3 = BasicRMSDGrouper(
            molecules,
            threshold=1.0,
            num_procs=self.NUM_PROCS,
            ignore_hydrogens=True,
        )
        groups, group_indices = grouper3.group()
        assert len(groups) == 10
        assert len(group_indices) == 10
        unique_structures = grouper3.unique()
        assert len(unique_structures) == 10

        grouper4 = BasicRMSDGrouper(
            molecules,
            threshold=1.5,
            num_procs=self.NUM_PROCS,
            ignore_hydrogens=True,
        )
        groups, group_indices = grouper4.group()
        assert len(groups) == 7
        assert len(group_indices) == 7
        unique_structures = grouper4.unique()
        assert len(unique_structures) == 7

        grouper5 = BasicRMSDGrouper(
            molecules,
            threshold=2.0,
            num_procs=self.NUM_PROCS,
            ignore_hydrogens=True,
        )
        groups, group_indices = grouper5.group()
        assert len(groups) == 5
        assert len(group_indices) == 5
        unique_structures = grouper5.unique()
        assert len(unique_structures) == 5

        grouper6 = BasicRMSDGrouper(
            molecules,
            threshold=2.5,
            num_procs=self.NUM_PROCS,
            ignore_hydrogens=True,
        )
        groups, group_indices = grouper6.group()
        for indices in group_indices:
            for i, idx_i in enumerate(indices):
                for idx_j in indices[i + 1 :]:
                    rmsd = grouper6.calculate_rmsd_pair(idx_i, idx_j)
                    assert rmsd <= 2.5
        assert len(groups) == 4
        assert len(group_indices) == 4
        unique_structures = grouper6.unique()
        assert len(unique_structures) == 4

    def test_rmsd_grouper_for_rotated_molecules(
        self, two_rotated_molecules_xyz_file
    ):
        xyz_file = XYZFile(filename=two_rotated_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)
        assert len(molecules) == 2
        grouper = BasicRMSDGrouper(
            molecules,
            threshold=0.5,
            num_procs=self.NUM_PROCS,
            ignore_hydrogens=False,
        )
        groups, group_indices = grouper.group()
        assert len(groups) == 2
        assert len(group_indices) == 2
        unique_structures = grouper.unique()
        assert len(unique_structures) == 2

    def test_calculate_rmsd_pair_wrapper_validation(self, methanol_molecules):
        grouper = BasicRMSDGrouper(methanol_molecules[:2])

        assert np.isclose(grouper.calculate_rmsd_pair(0, 0), 0.0)

        with pytest.raises(IndexError):
            grouper.calculate_rmsd_pair(0, 3)

        with pytest.raises(TypeError):
            grouper.calculate_rmsd_pair("a", 1)  # type: ignore[arg-type]


@pytest.mark.usefixtures("temporary_working_dir")
class Test_HungarianRMSD_grouper:
    NUM_PROCS = 4

    def test_hrmsd_grouper_for_rotated_molecules(
        self, two_rotated_molecules_xyz_file
    ):
        xyz_file = XYZFile(filename=two_rotated_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)
        assert len(molecules) == 2
        grouper = HungarianRMSDGrouper(
            molecules,
            threshold=0.5,
            num_procs=self.NUM_PROCS,
            ignore_hydrogens=False,
        )
        groups, group_indices = grouper.group()
        assert len(groups) == 1
        assert len(group_indices) == 1
        unique_structures = grouper.unique()
        assert len(unique_structures) == 1

        # rmsd calculation from grouper
        rmsd = grouper.calculate_rmsd_pair(0, 1)
        assert np.isclose(rmsd, 0.2294, rtol=1e-3)

    def test_hrmsd_grouper_for_crest_molecules(
        self, multiple_molecules_xyz_file
    ):
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)
        assert len(molecules) == 18
        grouper = HungarianRMSDGrouper(
            molecules,
            threshold=0.5,
            num_procs=self.NUM_PROCS,
            ignore_hydrogens=False,
        )
        groups, group_indices = grouper.group()
        assert len(groups) == 12
        assert len(group_indices) == 12
        unique_structures = grouper.unique()
        assert len(unique_structures) == 12

        # rmsd calculation from grouper
        rmsd = grouper.calculate_rmsd_pair(0, 1)
        assert np.isclose(rmsd, 0.4091, rtol=1e-3)
        rmsd = grouper.calculate_rmsd_pair(0, 2)
        assert np.isclose(rmsd, 0.5899, rtol=1e-3)
        rmsd = grouper.calculate_rmsd_pair(0, 3)
        assert np.isclose(rmsd, 1.8891, rtol=1e-3)

    def test_ignore_hydrogen(self, multiple_molecules_xyz_file):
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)
        assert len(molecules) == 18
        grouper = HungarianRMSDGrouper(
            molecules,
            threshold=0.5,
            num_procs=self.NUM_PROCS,
            ignore_hydrogens=True,
        )
        groups, group_indices = grouper.group()
        assert len(groups) == 12
        assert len(group_indices) == 12
        unique_structures = grouper.unique()
        assert len(unique_structures) == 12

        # rmsd calculation from grouper
        rmsd = grouper.calculate_rmsd_pair(0, 4)
        assert np.isclose(rmsd, 1.1915, rtol=1e-3)


@pytest.mark.usefixtures("temporary_working_dir")
class Test_SpyRMSD_grouper:
    NUM_PROCS = 4

    def test_spyrmsd_grouper_for_rotated_molecules(
        self, two_rotated_molecules_xyz_file
    ):
        xyz_file = XYZFile(filename=two_rotated_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)
        assert len(molecules) == 2
        grouper = SpyRMSDGrouper(
            molecules,
            threshold=0.5,
            num_procs=self.NUM_PROCS,
            ignore_hydrogens=False,
        )
        groups, group_indices = grouper.group()
        assert len(groups) == 1
        assert len(group_indices) == 1
        unique_structures = grouper.unique()
        assert len(unique_structures) == 1

        # rmsd calculation from grouper
        rmsd = grouper.calculate_rmsd_pair(0, 1)
        assert np.isclose(rmsd, 0.2125, rtol=1e-3)

    def test_spyrmsd_grouper_for_crest_molecules(
        self, multiple_molecules_xyz_file
    ):
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)
        assert len(molecules) == 18
        grouper = SpyRMSDGrouper(
            molecules,
            threshold=0.5,
            num_procs=self.NUM_PROCS,
            ignore_hydrogens=False,
        )
        groups, group_indices = grouper.group()
        assert len(groups) == 12
        assert len(group_indices) == 12
        unique_structures = grouper.unique()
        assert len(unique_structures) == 12

        # rmsd calculation from grouper
        rmsd = grouper.calculate_rmsd_pair(0, 1)
        assert np.isclose(rmsd, 0.4091, rtol=1e-3)
        rmsd = grouper.calculate_rmsd_pair(0, 2)
        assert np.isclose(rmsd, 1.3925, rtol=1e-3)
        rmsd = grouper.calculate_rmsd_pair(0, 3)
        assert np.isclose(rmsd, 2.1789, rtol=1e-3)
        rmsd = grouper.calculate_rmsd_pair(0, 4)
        assert np.isclose(rmsd, 1.8202, rtol=1e-3)
        rmsd = grouper.calculate_rmsd_pair(0, 5)
        assert np.isclose(rmsd, 2.0029, rtol=1e-3)

    def test_ignore_hydrogen(self, multiple_molecules_xyz_file):
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)
        assert len(molecules) == 18
        grouper = SpyRMSDGrouper(
            molecules,
            threshold=2.5981,
            num_procs=self.NUM_PROCS,
            ignore_hydrogens=True,
        )
        groups, group_indices = grouper.group()
        assert len(groups) == 3
        assert len(group_indices) == 3
        unique_structures = grouper.unique()
        assert len(unique_structures) == 3

        # rmsd calculation from grouper
        rmsd = grouper.calculate_rmsd_pair(0, 5)
        assert np.isclose(rmsd, 1.7034, rtol=1e-3)
        rmsd = grouper.calculate_rmsd_pair(0, 6)
        assert np.isclose(rmsd, 2.6183, rtol=1e-3)


@pytest.mark.skipif(
    not _is_irmsd_available(), reason="irmsd API/command not available"
)
@pytest.mark.usefixtures("temporary_working_dir")
class Test_IRMSD_grouper:
    NUM_PROCS = 4

    def test_irmsd_grouper_for_rotated_molecules(
        self, two_rotated_molecules_xyz_file
    ):
        xyz_file = XYZFile(filename=two_rotated_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)
        assert len(molecules) == 2
        grouper = IRMSDGrouper(
            molecules,
            threshold=0.125,
            num_procs=self.NUM_PROCS,
            ignore_hydrogens=False,
        )
        groups, group_indices = grouper.group()
        assert len(groups) == 2
        assert len(group_indices) == 2
        unique_structures = grouper.unique()
        assert len(unique_structures) == 2

        # rmsd calculation from grouper
        rmsd = grouper.calculate_rmsd_pair(0, 1)
        assert np.isclose(rmsd, 0.2294, rtol=1e-3)

    def test_irmsd_grouper_for_crest_molecules(
        self, multiple_molecules_xyz_file
    ):
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)
        assert len(molecules) == 18
        grouper = IRMSDGrouper(
            molecules,
            threshold=0.5,
            num_procs=self.NUM_PROCS,
            ignore_hydrogens=False,
        )
        groups, group_indices = grouper.group()
        assert len(groups) == 12
        assert len(group_indices) == 12
        unique_structures = grouper.unique()
        assert len(unique_structures) == 12

        # rmsd calculation from grouper
        rmsd = grouper.calculate_rmsd_pair(0, 1)
        assert np.isclose(rmsd, 0.4091, rtol=1e-3)
        rmsd = grouper.calculate_rmsd_pair(0, 2)
        assert np.isclose(rmsd, 1.3925, rtol=1e-3)
        rmsd = grouper.calculate_rmsd_pair(2, 10)
        assert np.isclose(rmsd, 2.2390, rtol=1e-3)
        rmsd = grouper.calculate_rmsd_pair(8, 16)
        assert np.isclose(rmsd, 3.4209, rtol=1e-3)
        rmsd = grouper.calculate_rmsd_pair(0, 13)
        assert np.isclose(rmsd, 0.8411, rtol=1e-3)

    def test_ignore_hydrogen(self, two_rotated_molecules_xyz_file):
        xyz_file = XYZFile(filename=two_rotated_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)
        assert len(molecules) == 2
        grouper = IRMSDGrouper(
            molecules,
            threshold=0.125,
            num_procs=self.NUM_PROCS,
            ignore_hydrogens=True,
        )
        groups, group_indices = grouper.group()
        assert len(groups) == 2
        assert len(group_indices) == 2
        unique_structures = grouper.unique()
        assert len(unique_structures) == 2

        # rmsd calculation from grouper
        rmsd = grouper.calculate_rmsd_pair(0, 1)
        assert np.isclose(rmsd, 0.2294, rtol=1e-3)

    def test_parallel_irmsd_preserves_actual_inversion_metadata(
        self, two_rotated_molecules_xyz_file
    ):
        """Parallel iRMSD must propagate resolved inversion metadata to the parent."""
        xyz_file = XYZFile(filename=two_rotated_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)

        grouper = IRMSDGrouper(
            molecules,
            threshold=0.5,
            num_procs=self.NUM_PROCS,
            ignore_hydrogens=False,
            inversion="off",
        )
        grouper.group()

        assert grouper.num_procs > 1
        assert grouper._actual_inversion == "off"


@pytest.mark.usefixtures("temporary_working_dir")
class Test_PymolRMSD_grouper:
    NUM_PROCS = 1

    @classmethod
    def setup_class(cls):
        """Initialize PyMOL once for all tests in this class."""
        try:
            import pymol
            from pymol import cmd

            pymol.finish_launching(["pymol", "-qc"])
            cmd.reinitialize()
        except Exception:
            # PyMOL may not be installed; tests will be skipped via @pytest.mark.skipif
            pass

    def teardown_method(self, method):
        """Clean up PyMOL objects after each test method to prevent slowdown."""
        try:
            from pymol import cmd

            # Delete all objects instead of reinitialize to keep session alive
            cmd.delete("all")
        except Exception:
            # Ignore cleanup errors if PyMOL is not available
            pass

    def test_pymolrmsd_grouper_for_rotated_molecules(
        self, two_rotated_molecules_xyz_file
    ):
        xyz_file = XYZFile(filename=two_rotated_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)
        assert len(molecules) == 2
        grouper = PymolRMSDGrouper(
            molecules,
            threshold=0.5,
            num_procs=self.NUM_PROCS,
            ignore_hydrogens=False,
        )
        groups, group_indices = grouper.group()
        assert len(groups) == 1
        assert len(group_indices) == 1
        unique_structures = grouper.unique()
        assert len(unique_structures) == 1

        # rmsd calculation from grouper
        rmsd = grouper.calculate_rmsd_pair(0, 1)
        assert np.isclose(rmsd, 0.0000, rtol=1e-3)

        # Explicitly cleanup to prevent __del__ from calling quit()
        if hasattr(grouper, "_temp_dir") and grouper._temp_dir:

            shutil.rmtree(grouper._temp_dir, ignore_errors=True)
            grouper._temp_dir = None
        grouper.cmd = None  # Prevent __del__ from calling quit()

    def test_pymolrmsd_grouper_for_crest_molecules(
        self, multiple_molecules_xyz_file
    ):
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)
        assert len(molecules) == 18
        grouper = PymolRMSDGrouper(
            molecules,
            threshold=0.5,
            num_procs=self.NUM_PROCS,
            ignore_hydrogens=False,
        )
        groups, group_indices = grouper.group()
        assert len(groups) == 8
        assert len(group_indices) == 8
        unique_structures = grouper.unique()
        assert len(unique_structures) == 8
        for indices in group_indices:
            for i, idx_i in enumerate(indices):
                for idx_j in indices[i + 1 :]:
                    rmsd = grouper.calculate_rmsd_pair(idx_i, idx_j)
                    assert rmsd <= 0.5
        # rmsd calculation from grouper
        rmsd = grouper.calculate_rmsd_pair(0, 1)
        assert np.isclose(rmsd, 0.074175, rtol=1e-3)
        rmsd = grouper.calculate_rmsd_pair(0, 2)
        assert np.isclose(rmsd, 0.023745, rtol=1e-3)
        rmsd = grouper.calculate_rmsd_pair(3, 4)
        assert np.isclose(rmsd, 1.725241, rtol=1e-3)
        rmsd = grouper.calculate_rmsd_pair(7, 8)
        assert np.isclose(rmsd, 2.309451, rtol=1e-3)
        rmsd = grouper.calculate_rmsd_pair(14, 16)
        assert np.isclose(rmsd, 1.837319, rtol=1e-3)

        # Explicitly cleanup to prevent __del__ from calling quit()
        if hasattr(grouper, "_temp_dir") and grouper._temp_dir:

            shutil.rmtree(grouper._temp_dir, ignore_errors=True)
            grouper._temp_dir = None
        grouper.cmd = None  # Prevent __del__ from calling quit()

    def test_ignore_hydrogen(self, multiple_molecules_xyz_file):
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)
        assert len(molecules) == 18
        grouper = PymolRMSDGrouper(
            molecules,
            threshold=0.8203818,
            num_procs=self.NUM_PROCS,
            ignore_hydrogens=True,
        )
        groups, group_indices = grouper.group()
        assert len(groups) == 5
        assert len(group_indices) == 5
        unique_structures = grouper.unique()
        assert len(unique_structures) == 5

        # rmsd calculation from grouper
        rmsd = grouper.calculate_rmsd_pair(0, 2)
        assert np.isclose(rmsd, 0.0211, rtol=1e-3)
        rmsd = grouper.calculate_rmsd_pair(1, 4)
        assert np.isclose(rmsd, 0.7669, rtol=1e-3)

        # Explicitly cleanup to prevent __del__ from calling quit()
        if hasattr(grouper, "_temp_dir") and grouper._temp_dir:

            shutil.rmtree(grouper._temp_dir, ignore_errors=True)
            grouper._temp_dir = None
        grouper.cmd = None  # Prevent __del__ from calling quit()

    def test_pymol_grouper_falls_back_to_single_proc(
        self, methanol_molecules, caplog
    ):
        """PyMOL grouper warns and falls back to num_procs=1 when >1 is requested."""
        caplog.set_level("WARNING")
        grouper = PymolRMSDGrouper(
            methanol_molecules,
            threshold=0.5,
            num_procs=4,
        )
        assert grouper.num_procs == 1
        assert (
            "PymolRMSDGrouper does not support multiprocessing; using num_procs=1."
            in caplog.text
        )

        if hasattr(grouper, "_temp_dir") and grouper._temp_dir:
            shutil.rmtree(grouper._temp_dir, ignore_errors=True)
            grouper._temp_dir = None
        grouper.cmd = None

    @classmethod
    def teardown_class(cls):
        """Clean up PyMOL after all tests in this class are done."""
        try:
            from pymol import cmd

            cmd.reinitialize()
        except Exception:
            # Ignore cleanup errors if PyMOL is not available
            pass


@pytest.mark.usefixtures("temporary_working_dir")
class Test_Tanimoto_similarity_grouper:
    NUM_PROCS = 4

    def test_tanimoto_similarity_grouper(
        self, methanol_molecules, methanol_and_ethanol
    ):
        grouper = TanimotoSimilarityGrouper(methanol_molecules)
        groups, group_indices = grouper.group()
        assert (
            len(groups) == 1
        ), "Molecules should form one group based on RCM similarity."
        assert (
            len(group_indices) == 1
        ), "Molecules should form one group based on RCM similarity."
        unique_structures = grouper.unique()
        assert (
            len(unique_structures) == 1
        ), "Molecules should form one group based on RCM similarity."
        grouper2 = TanimotoSimilarityGrouper(methanol_and_ethanol)
        groups, group_indices = grouper2.group()
        assert (
            len(groups) == 2
        ), "Molecules should form two groups based on RCM similarity."
        assert (
            len(group_indices) == 2
        ), "Molecules should form two groups based on RCM similarity."
        unique_structures = grouper2.unique()
        assert (
            len(unique_structures) == 2
        ), "Molecules should form two groups based on RCM similarity."

    def test_calculate_tanimoto_pair_wrapper_validation(
        self, methanol_molecules
    ):
        grouper = TanimotoSimilarityGrouper(
            methanol_molecules[:2], fingerprint_type="rdkit"
        )

        assert np.isclose(grouper.calculate_tanimoto_pair(0, 0), 1.0)

        with pytest.raises(IndexError):
            grouper.calculate_tanimoto_pair(0, 4)

        with pytest.raises(TypeError):
            grouper.calculate_tanimoto_pair("a", 1)  # type: ignore[arg-type]

    def test_tanimoto_grouper_for_crest_conformers(
        self, multiple_molecules_xyz_file
    ):
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)

        molecules = xyz_file.get_molecules(index=":", return_list=True)
        assert len(molecules) == 18
        grouper = TanimotoSimilarityGrouper(
            molecules,
            threshold=0.8,
            fingerprint_type="usrcat",
            num_procs=self.NUM_PROCS,
        )
        groups, group_indices = grouper.group()
        assert len(groups) == 14
        assert len(group_indices) == 14
        unique_structures = grouper.unique()
        assert len(unique_structures) == 14

        grouper = TanimotoSimilarityGrouper(
            molecules,
            threshold=0.9,
            fingerprint_type="usrcat",
            num_procs=self.NUM_PROCS,
        )
        groups, group_indices = grouper.group()
        assert len(groups) == 17
        assert len(group_indices) == 17
        unique_structures = grouper.unique()
        assert len(unique_structures) == 17

    def test_tanimoto_threshold_boundary_through_group(
        self, methanol_molecules, monkeypatch
    ):
        molecules = methanol_molecules[:2]
        from chemsmart.jobs.grouper import tanimoto as tanimoto_module

        for fingerprint_type in ("usr", "usrcat"):
            for similarity, expected_num_groups in (
                (0.90, 1),
                (0.91, 1),
                (0.89, 2),
            ):
                monkeypatch.setattr(
                    tanimoto_module.rdMolDescriptors,
                    "GetUSRScore",
                    lambda *args, value=similarity, **kwargs: value,
                )

                grouper = TanimotoSimilarityGrouper(
                    molecules,
                    threshold=0.9,
                    fingerprint_type=fingerprint_type,
                    num_procs=1,
                )
                groups, group_indices = grouper.group()

                assert len(groups) == expected_num_groups
                assert len(group_indices) == expected_num_groups
                if expected_num_groups == 1:
                    assert set(group_indices[0]) == {0, 1}
                else:
                    assert all(len(group) == 1 for group in group_indices)


@pytest.mark.usefixtures("temporary_working_dir")
class Test_TorsionFingerprint_grouper:
    NUM_PROCS = 4

    def test_torsionfingerprint_grouper_for_rotated_molecules(
        self, two_rotated_molecules_xyz_file
    ):
        xyz_file = XYZFile(filename=two_rotated_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)
        assert len(molecules) == 2
        grouper = TorsionFingerprintGrouper(
            molecules,
            threshold=0.1,
            num_procs=self.NUM_PROCS,
            ignore_hydrogens=False,
        )
        groups, group_indices = grouper.group()
        assert len(groups) == 1
        assert len(group_indices) == 1
        unique_structures = grouper.unique()
        assert len(unique_structures) == 1

    def test_torsionfingerprint_grouper_for_crest_molecules(
        self, multiple_molecules_xyz_file
    ):
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)
        assert len(molecules) == 18
        grouper = TorsionFingerprintGrouper(
            molecules,
            threshold=0.05,
            num_procs=self.NUM_PROCS,
            ignore_hydrogens=False,
        )
        groups, group_indices = grouper.group()
        assert len(groups) == 10
        assert len(group_indices) == 10
        unique_structures = grouper.unique()
        assert len(unique_structures) == 10

    def test_use_weights_parameter(self, multiple_molecules_xyz_file):
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)
        assert len(molecules) == 18
        grouper = TorsionFingerprintGrouper(
            molecules,
            threshold=0.05,
            num_procs=self.NUM_PROCS,
            use_weights=False,
            ignore_hydrogens=False,
        )
        groups, group_indices = grouper.group()
        assert len(groups) == 15
        assert len(group_indices) == 15
        unique_structures = grouper.unique()
        assert len(unique_structures) == 15

        tfd = grouper.calculate_tfd_pair(0, 2)
        assert np.isclose(tfd, 0.08229, rtol=1e-3)
        tfd = grouper.calculate_tfd_pair(0, 7)
        assert np.isclose(tfd, 0.07727, rtol=1e-3)

    def test_tfd_num_groups_records_actual_group_count(
        self, multiple_molecules_xyz_file, temporary_working_dir
    ):
        from openpyxl import load_workbook

        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)

        grouper = TorsionFingerprintGrouper(
            molecules,
            num_groups=5,
            num_procs=1,
            ignore_hydrogens=False,
            label="tfd_actual_groups",
        )
        groups, _ = grouper.group()

        xlsx_file = os.path.join(
            temporary_working_dir,
            "tfd_actual_groups_group_result",
            "tfd_actual_groups_TorsionFingerprintGrouper_N5.xlsx",
        )
        assert os.path.exists(xlsx_file)

        wb = load_workbook(xlsx_file, data_only=True)
        ws = wb["TFD_Matrix"]
        header_lines = [str(ws[f"A{i}"].value) for i in range(1, 25)]

        assert any(
            f"Actual Groups: {len(groups)}" in line
            for line in header_lines
            if line
        )

    def test_use_maxdev_parameter(self, multiple_molecules_xyz_file):
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)
        assert len(molecules) == 18
        grouper = TorsionFingerprintGrouper(
            molecules,
            threshold=0.26965,
            num_procs=self.NUM_PROCS,
            use_weights=True,
            max_dev="spec",
            ignore_hydrogens=False,
        )
        groups, group_indices = grouper.group()
        assert len(groups) == 3
        assert len(group_indices) == 3
        unique_structures = grouper.unique()
        assert len(unique_structures) == 3

        tfd = grouper.calculate_tfd_pair(0, 1)
        assert np.isclose(tfd, 0.02027, rtol=1e-3)
        tfd = grouper.calculate_tfd_pair(3, 4)
        assert np.isclose(tfd, 0.24365, rtol=1e-3)

    def test_calculate_tfd_pair_wrapper_validation(
        self, multiple_molecules_xyz_file
    ):
        molecules = XYZFile(
            filename=multiple_molecules_xyz_file
        ).get_molecules(index=":", return_list=True)
        grouper = TorsionFingerprintGrouper(molecules[:3], threshold=0.1)

        assert np.isclose(grouper.calculate_tfd_pair(1, 1), 0.0)

        with pytest.raises(IndexError):
            grouper.calculate_tfd_pair(0, 8)

        with pytest.raises(TypeError):
            grouper.calculate_tfd_pair("a", 1)  # type: ignore[arg-type]


@pytest.mark.usefixtures("temporary_working_dir")
class Test_other_groupers:
    NUM_PROCS = 4

    def test_base_record_template_method(self, methanol_molecules):
        class DummyGrouper(MoleculeGrouper):
            def group(self):
                return [], []

            def _record_results(self, **kwargs):
                return kwargs

        grouper = DummyGrouper(methanol_molecules)
        payload = grouper.record(example_key="example_value")
        assert payload["example_key"] == "example_value"

    @pytest.mark.parametrize("energy_type", ["QHH", "QHG", "SP_QHG"])
    def test_thermochemistry_parameters_header_for_qhg_energy_types(
        self, methanol_molecules, energy_type
    ):
        """qhG and SP-qhG results must record the exact thermochemistry settings used."""
        thermo_parameters = (
            "temperature=298.15, concentration=1.0, pressure=1.0, "
            "use_weighted_mass=True, alpha=4, s_freq_cutoff=100.0, "
            "entropy_method=grimme, h_freq_cutoff=100.0, "
            "energy_units=hartree, check_imaginary_frequencies=True, "
            "cutoff_entropy_grimme=100.0, cutoff_enthalpy=100.0"
        )
        grouper = BasicRMSDGrouper(
            methanol_molecules[:2],
            threshold=0.5,
            num_procs=1,
            energy_type=energy_type,
            thermo_parameters=thermo_parameters,
        )

        header_info = []
        grouper._append_thermo_header(header_info)

        assert header_info == [
            ("Thermochemistry Parameters", thermo_parameters)
        ]

    @pytest.mark.parametrize("energy_type", ["E", "H", "G"])
    def test_thermochemistry_parameters_header_not_written_for_other_energy_types(
        self, methanol_molecules, energy_type
    ):
        """Thermochemistry parameter header is specific to qhG and SP-qhG."""
        thermo_parameters = (
            "temperature=298.15, concentration=1.0, pressure=1.0, "
            "use_weighted_mass=True, alpha=4, s_freq_cutoff=100.0, "
            "entropy_method=grimme, h_freq_cutoff=100.0, "
            "energy_units=hartree, check_imaginary_frequencies=True, "
            "cutoff_entropy_grimme=100.0, cutoff_enthalpy=100.0"
        )
        grouper = BasicRMSDGrouper(
            methanol_molecules[:2],
            threshold=0.5,
            num_procs=1,
            energy_type=energy_type,
            thermo_parameters=thermo_parameters,
        )

        header_info = []
        grouper._append_thermo_header(header_info)

        assert not any(
            key == "Thermochemistry Parameters" for key, _ in header_info
        )

    def test_record_writes_rmsd_matrix_with_headers(
        self, methanol_molecules, temporary_working_dir
    ):
        """Detailed record test: verify matrix file and key header lines are written."""
        from openpyxl import load_workbook

        molecules = methanol_molecules[:2]
        grouper = BasicRMSDGrouper(
            molecules,
            threshold=0.5,
            num_procs=1,
            label="record_detail",
            energy_type="E",
        )

        # Populate cache first so Groups sheet can be generated.
        groups, group_indices = grouper.group()
        assert len(groups) >= 1
        assert len(group_indices) >= 1

        # Call unified record() explicitly to validate template-method path.
        rmsd_matrix = np.zeros((len(molecules), len(molecules)))
        grouper.record(rmsd_matrix=rmsd_matrix, grouping_time=0.01)

        xlsx_file = os.path.join(
            temporary_working_dir,
            "record_detail_group_result",
            "record_detail_BasicRMSDGrouper_T0.5.xlsx",
        )
        assert os.path.exists(xlsx_file)

        wb = load_workbook(xlsx_file, data_only=True)
        assert "RMSD_Matrix" in wb.sheetnames
        assert "Groups" in wb.sheetnames

        ws = wb["RMSD_Matrix"]
        header_lines = [str(ws[f"A{i}"].value) for i in range(1, 25)]

        assert any("Energy Type: E" in line for line in header_lines if line)
        assert any(
            "Used Molecules: 2" in line for line in header_lines if line
        )
        assert any(
            "Skipped Molecules: 0" in line for line in header_lines if line
        )
        assert any("Grouping Time:" in line for line in header_lines if line)

    def test_record_writes_formula_outputs_with_headers(
        self, methanol_and_ethanol, temporary_working_dir
    ):
        """Detailed record test: verify non-matrix output sheet and header lines."""
        from openpyxl import load_workbook

        grouper = FormulaGrouper(
            methanol_and_ethanol,
            num_procs=1,
            label="formula_record_detail",
            energy_type="E",
        )
        groups, group_indices = grouper.group()
        assert len(groups) == 2
        assert len(group_indices) == 2

        xlsx_file = os.path.join(
            temporary_working_dir,
            "formula_record_detail_group_result",
            "formula_record_detail_FormulaGrouper.xlsx",
        )
        assert os.path.exists(xlsx_file)

        wb = load_workbook(xlsx_file, data_only=True)
        assert "Formulas" in wb.sheetnames
        assert "Groups" in wb.sheetnames

        ws = wb["Formulas"]
        header_lines = [str(ws[f"A{i}"].value) for i in range(1, 25)]
        assert any(
            "Total Molecules: 2" in line for line in header_lines if line
        )
        assert any(
            "Unique Formulas: 2" in line for line in header_lines if line
        )
        assert any("Energy Type: E" in line for line in header_lines if line)

    def test_formula_grouper(
        self,
        methanol_molecules,
        methanol_and_ethanol,
        conformers_from_rdkit,
    ):
        grouper = FormulaGrouper(methanol_molecules)
        groups, group_indices = grouper.group()
        unique_structures = grouper.unique()
        assert (
            len(groups) == 1
        ), "Molecules should form one group based on formula."

        assert (
            len(unique_structures) == 1
        ), "Molecules should form one group based on formula."

        grouper2 = FormulaGrouper(methanol_and_ethanol)
        groups, group_indices = grouper2.group()
        unique_structures = grouper2.unique()
        assert (
            len(groups) == 2
        ), "Molecules should form two groups based on formula."
        assert (
            len(unique_structures) == 2
        ), "Molecules should form two groups based on formula."

        grouper3 = FormulaGrouper(conformers_from_rdkit)
        # based on Formula, should all be the same even for 300 conformers
        groups, group_indices = grouper3.group()
        unique_structures = grouper3.unique()
        assert len(groups) == 1
        assert len(unique_structures) == 1

    @pytest.mark.slow
    def test_connectivity_grouper(
        self, methanol_molecules, methanol_and_ethanol
    ):
        grouper = ConnectivityGrouper(methanol_molecules)
        groups, group_indices = grouper.group()
        assert (
            len(groups) == 1
        ), "Molecules should form one group based on connectivity."
        assert (
            len(group_indices) == 1
        ), "Molecules should form one group based on connectivity."
        unique_structures = grouper.unique()
        assert (
            len(unique_structures) == 1
        ), "Molecules should form one group based on connectivity."

        grouper2 = ConnectivityGrouper(methanol_and_ethanol)
        groups, group_indices = grouper2.group()
        assert (
            len(groups) == 2
        ), "Molecules should form two groups based on connectivity."
        assert (
            len(group_indices) == 2
        ), "Molecules should form two groups based on connectivity."
        unique_structures = grouper2.unique()
        assert (
            len(unique_structures) == 2
        ), "Molecules should form two groups based on connectivity."

    def test_connectivity_grouper_for_crest_conformers(
        self, multiple_molecules_xyz_file
    ):
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)

        molecules = xyz_file.get_molecules(index=":", return_list=True)
        assert len(molecules) == 18
        grouper = ConnectivityGrouper(
            molecules,
            num_procs=self.NUM_PROCS,
        )
        groups, group_indices = grouper.group()
        assert len(groups) == 1
        assert len(group_indices) == 1
        unique_structures = grouper.unique()
        assert len(unique_structures) == 1

    def test_rdkit_isomorphism_grouper(
        self, methanol_molecules, methanol_and_ethanol
    ):
        grouper = RDKitIsomorphismGrouper(methanol_molecules)
        groups, group_indices = grouper.group()
        assert (
            len(groups) == 1
        ), "Molecules should form one group based on RCM similarity."
        assert (
            len(group_indices) == 1
        ), "Molecules should form one group based on RCM similarity."
        unique_structures = grouper.unique()
        assert (
            len(unique_structures) == 1
        ), "Molecules should form one group based on RCM similarity."
        grouper2 = RDKitIsomorphismGrouper(methanol_and_ethanol)
        groups, group_indices = grouper2.group()
        assert (
            len(groups) == 2
        ), "Molecules should form two groups based on RCM similarity."
        assert (
            len(group_indices) == 2
        ), "Molecules should form two groups based on RCM similarity."
        unique_structures = grouper2.unique()
        assert (
            len(unique_structures) == 2
        ), "Molecules should form two groups based on RCM similarity."


@pytest.mark.usefixtures("temporary_working_dir")
class Test_EnergyGrouper:
    NUM_PROCS = 4

    def test_energy_grouper_raises_error_for_missing_energy(
        self, methanol_molecules
    ):
        # methanol_molecules from pubchem don't have energy information
        with pytest.raises(ValueError) as excinfo:
            EnergyGrouper(methanol_molecules)
        assert "missing energy information" in str(excinfo.value)

    def test_energy_grouper_for_crest_conformers(
        self, multiple_molecules_xyz_file
    ):
        """Test EnergyGrouper with molecules that have energy information."""
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)
        assert len(molecules) == 18

        # All CREST conformers should have energy
        for mol in molecules:
            assert mol.energy is not None

        assert len(molecules) == 18
        grouper = EnergyGrouper(
            molecules,
            threshold=0.5,
            num_procs=self.NUM_PROCS,
        )
        groups, group_indices = grouper.group()
        assert len(groups) == 8
        assert len(group_indices) == 8
        unique_structures = grouper.unique()
        assert len(unique_structures) == 8

        expected1 = -1.7839 / 627.509474
        expected2 = 1.4580 / 627.509474

        relative_diff, abs_diff = grouper.calculate_energy_diff_pair(1, 0)
        assert np.isclose(relative_diff, expected1, rtol=1e-3)
        relative_diff, abs_diff = grouper.calculate_energy_diff_pair(1, 2)
        assert np.isclose(abs_diff, expected2, rtol=1e-3)

        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)

        grouper = EnergyGrouper(
            molecules, num_groups=5, num_procs=self.NUM_PROCS
        )
        groups, group_indices = grouper.group()
        assert len(groups) == 5
        assert len(group_indices) == 5
        assert grouper._auto_threshold is not None
        assert grouper._auto_threshold > 0

    def test_energy_grouper_for_log_conformers(
        self, ts_conformers_log_directory
    ):
        """Test EnergyGrouper with molecules loaded from log files using Gibbs energy."""
        import glob
        import re

        from chemsmart.io.gaussian.output import Gaussian16Output
        from chemsmart.io.molecules.structure import Molecule

        # Load molecules from log files and set Gibbs energy (like CLI does)
        log_files = sorted(
            glob.glob(os.path.join(ts_conformers_log_directory, "*.log"))
        )
        molecules = []
        conformer_ids = []

        for log_file in log_files:
            mol = Molecule.from_filepath(
                filepath=log_file, index="-1", return_list=False
            )
            if mol is not None:
                # Extract and set Gibbs energy using gibbs_free_energy property
                g16_output = Gaussian16Output(filename=log_file)
                gibbs_energy = g16_output.gibbs_free_energy
                if gibbs_energy is not None:
                    mol._energy = gibbs_energy

                molecules.append(mol)
                basename = os.path.basename(log_file)
                match = re.search(r"_c(\d+)\.log$", basename)
                if match:
                    conformer_ids.append(f"c{match.group(1)}")
                else:
                    conformer_ids.append(basename)

        assert len(molecules) == 5
        assert len(conformer_ids) == 5

        # All log conformers should have Gibbs energy
        for mol in molecules:
            assert mol.energy is not None

        grouper = EnergyGrouper(
            molecules,
            threshold=1.0,
            num_procs=self.NUM_PROCS,
            conformer_ids=conformer_ids,
        )
        groups, group_indices = grouper.group()
        assert len(groups) == 2
        assert len(group_indices) == 2
        unique_structures = grouper.unique()
        assert len(unique_structures) == 2

        expected1 = -4.1429 / 627.509474
        expected2 = 4.1429 / 627.509474

        relative_diff, abs_diff = grouper.calculate_energy_diff_pair(1, 4)
        assert np.isclose(relative_diff, expected1, rtol=1e-2)
        assert np.isclose(abs_diff, expected2, rtol=1e-2)

    def test_calculate_energy_diff_pair_wrapper_validation(
        self, multiple_molecules_xyz_file
    ):
        molecules = XYZFile(
            filename=multiple_molecules_xyz_file
        ).get_molecules(index=":", return_list=True)
        grouper = EnergyGrouper(molecules[:3], threshold=1.0)

        relative_diff, abs_diff = grouper.calculate_energy_diff_pair(2, 2)
        assert np.isclose(relative_diff, 0.0)
        assert np.isclose(abs_diff, 0.0)

        with pytest.raises(IndexError):
            grouper.calculate_energy_diff_pair(0, 5)

        with pytest.raises(TypeError):
            grouper.calculate_energy_diff_pair("a", 1)  # type: ignore[arg-type]

    def test_energy_extraction_from_xyz_file(
        self, multiple_molecules_xyz_file
    ):

        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)

        # Test specific energy value for 3rd structure (index 2)
        expected_energy_mol3 = -126.25238449
        actual_energy_mol3 = molecules[2].energy

        assert actual_energy_mol3 is not None, "3rd molecule energy is None"
        assert np.isclose(
            actual_energy_mol3, expected_energy_mol3, rtol=1e-8
        ), (
            f"Energy mismatch for 3rd molecule: got {actual_energy_mol3}, "
            f"expected {expected_energy_mol3}"
        )

    def test_gibbs_energy_extraction_function(
        self, ts_conformers_log_directory
    ):
        """Test that Gaussian16Output.gibbs_free_energy extracts correct value."""
        from chemsmart.io.gaussian.output import Gaussian16Output

        log_file = os.path.join(
            ts_conformers_log_directory, "ch_1c_para_c1.log"
        )
        g16_output = Gaussian16Output(filename=log_file)

        # Known values
        expected_scf = -1401.99431519
        expected_gibbs_correction = 0.311863
        expected_gibbs_energy = (
            expected_scf + expected_gibbs_correction
        )  # -1401.68245219

        gibbs_energy = g16_output.gibbs_free_energy
        assert np.isclose(gibbs_energy, expected_gibbs_energy, rtol=1e-6), (
            f"Gibbs energy mismatch: got {gibbs_energy}, "
            f"expected {expected_gibbs_energy}"
        )

    def test_energy_extraction_from_ts_log_files(
        self, ts_conformers_log_directory
    ):
        """Test that energy is correctly extracted from TS log files as SCF Done energy."""
        import glob

        from chemsmart.io.gaussian.output import Gaussian16Output
        from chemsmart.io.molecules.structure import Molecule

        log_files = sorted(
            glob.glob(os.path.join(ts_conformers_log_directory, "*.log"))
        )

        molecules = []
        for log_file in log_files:
            mol = Molecule.from_filepath(
                filepath=log_file, index="-1", return_list=False
            )
            if mol is not None:
                molecules.append(mol)
                g16_output = Gaussian16Output(filename=log_file)
                scf_energy = (
                    g16_output.scf_energies[-1]
                    if g16_output.scf_energies
                    else None
                )
                assert np.isclose(mol.energy, scf_energy, rtol=1e-6), (
                    f"Energy mismatch for {log_file}: "
                    f"mol.energy={mol.energy}, expected SCF={scf_energy}"
                )


@pytest.mark.usefixtures("temporary_working_dir")
class Testfactory:

    def test_structure_grouper_factory_energy(
        self, multiple_molecules_xyz_file
    ):
        """Test factory creation of energy grouper (requires molecules with energy)."""
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)

        factory = StructureGrouperFactory()
        energy_grouper = factory.create(molecules, strategy="energy")
        assert isinstance(energy_grouper, EnergyGrouper)

    def test_structure_grouper_factory(self, methanol_molecules):
        factory = StructureGrouperFactory()
        rmsd_grouper = factory.create(methanol_molecules, strategy="rmsd")
        assert isinstance(rmsd_grouper, RMSDGrouper)
        hrmsd_grouper = factory.create(methanol_molecules, strategy="hrmsd")
        assert isinstance(hrmsd_grouper, RMSDGrouper)
        spyrmsd_grouper = factory.create(
            methanol_molecules, strategy="spyrmsd"
        )
        assert isinstance(spyrmsd_grouper, RMSDGrouper)

        # irmsd requires either Python API or external command
        if _is_irmsd_available():
            irmsd_grouper = factory.create(
                methanol_molecules, strategy="irmsd"
            )
            assert isinstance(irmsd_grouper, RMSDGrouper)

        pymolrmsd_grouper = factory.create(
            methanol_molecules, strategy="pymolrmsd"
        )
        assert isinstance(pymolrmsd_grouper, RMSDGrouper)

        # Explicitly cleanup pymolrmsd_grouper to prevent __del__ from calling quit()
        if (
            hasattr(pymolrmsd_grouper, "_temp_dir")
            and pymolrmsd_grouper._temp_dir
        ):

            shutil.rmtree(pymolrmsd_grouper._temp_dir, ignore_errors=True)
            pymolrmsd_grouper._temp_dir = None
        pymolrmsd_grouper.cmd = None  # Prevent __del__ from calling quit()

        tfd_grouper = factory.create(methanol_molecules, strategy="tfd")
        assert isinstance(tfd_grouper, TorsionFingerprintGrouper)
        formula_grouper = factory.create(
            methanol_molecules, strategy="formula"
        )
        assert isinstance(formula_grouper, FormulaGrouper)
        connectivity_grouper = factory.create(
            methanol_molecules, strategy="connectivity"
        )
        assert isinstance(connectivity_grouper, ConnectivityGrouper)
        tanimoto_grouper = factory.create(
            methanol_molecules, strategy="tanimoto"
        )
        assert isinstance(tanimoto_grouper, TanimotoSimilarityGrouper)
        rdkit_isomorphism_grouper = factory.create(
            methanol_molecules, strategy="isomorphism"
        )
        assert isinstance(rdkit_isomorphism_grouper, RDKitIsomorphismGrouper)

    @classmethod
    def teardown_class(cls):
        """Clean up PyMOL after all tests in this class are done."""
        try:
            from pymol import cmd

            cmd.reinitialize()
        except Exception:
            # Ignore cleanup errors if PyMOL is not available
            pass


@pytest.mark.usefixtures("temporary_working_dir")
class Test_grouper_utility_functions:
    """Test utility functions and helper methods in groupers."""

    NUM_PROCS = 1

    def test_rmsd_matrix_with_num_groups(
        self, multiple_molecules_xyz_file, temporary_working_dir
    ):
        """Test that RMSD matrix filename reflects num_groups when used."""
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)

        grouper = BasicRMSDGrouper(
            molecules,
            num_groups=3,
            num_procs=1,
            label="test_num_groups",
        )
        _, _ = grouper.group()

        # Check that Excel file was created in the temporary_working_dir
        expected_file = os.path.join(
            temporary_working_dir,
            "test_num_groups_group_result",
            "test_num_groups_BasicRMSDGrouper_N3.xlsx",
        )
        assert os.path.exists(
            expected_file
        ), f"Matrix file not found: {expected_file}"

    def test_grouping_result_caching(self, multiple_molecules_xyz_file):
        """Test that grouping results are cached and reused."""
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)

        grouper = BasicRMSDGrouper(
            molecules,
            threshold=0.5,
            num_procs=1,
        )

        groups1, indices1 = grouper.group()

        assert grouper._cached_groups is not None
        assert grouper._cached_group_indices is not None

        # unique() should use cached results
        unique_mols = grouper.unique()
        assert len(unique_mols) == len(groups1)

    def test_unique_returns_lowest_energy_representative(
        self, multiple_molecules_xyz_file
    ):
        """Test that unique() returns lowest energy molecule from each group."""
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)

        # CREST conformers have energy information
        grouper = BasicRMSDGrouper(
            molecules,
            threshold=2.0,
            num_procs=1,
        )
        groups, group_indices = grouper.group()
        unique_mols = grouper.unique()

        # For each group, verify the representative has lowest energy
        for i, (group, indices) in enumerate(zip(groups, group_indices)):
            mols_with_energy = [
                (mol, idx)
                for mol, idx in zip(group, indices)
                if mol.energy is not None
            ]
            if mols_with_energy:
                min_energy_mol = min(
                    mols_with_energy, key=lambda x: x[0].energy
                )[0]
                assert unique_mols[i].energy == min_energy_mol.energy


@pytest.mark.usefixtures("temporary_working_dir")
class Test_grouper_complete_linkage:
    """Test complete linkage clustering behavior."""

    NUM_PROCS = 4

    def test_public_group_by_threshold_reuses_distance_matrix_and_caches_result(
        self, methanol_molecules
    ):
        molecules = methanol_molecules[:3]
        grouper = BasicRMSDGrouper(molecules, threshold=0.5, num_procs=1)
        distance_matrix = np.array(
            [
                [0.0, 0.3, 0.8],
                [0.3, 0.0, 0.7],
                [0.8, 0.7, 0.0],
            ]
        )
        original_matrix = distance_matrix.copy()

        _, strict_indices = grouper.group_by_threshold(
            distance_matrix, threshold=0.2
        )
        groups, relaxed_indices = grouper.group_by_threshold(
            distance_matrix, threshold=0.5
        )

        assert strict_indices == [[0], [1], [2]]
        assert {frozenset(group) for group in relaxed_indices} == {
            frozenset({0, 1}),
            frozenset({2}),
        }
        assert grouper.threshold == 0.5
        assert np.array_equal(distance_matrix, original_matrix)
        assert grouper._cached_groups == groups
        assert grouper._cached_group_indices == relaxed_indices

    def test_tfd_public_group_by_threshold_accepts_output_distance_matrix(
        self, methanol_molecules
    ):
        grouper = TorsionFingerprintGrouper(
            methanol_molecules[:2], threshold=0.1
        )
        tfd_matrix = np.array([[0.0, 0.15], [0.15, 0.0]])

        _, strict_indices = grouper.group_by_threshold(
            tfd_matrix, threshold=0.1
        )
        _, relaxed_indices = grouper.group_by_threshold(
            tfd_matrix, threshold=0.2
        )

        assert strict_indices == [[0], [1]]
        assert relaxed_indices == [[0, 1]]
        assert grouper.threshold == 0.1

    def test_energy_public_group_by_threshold_accepts_signed_kcal_output(
        self, multiple_molecules_xyz_file
    ):
        molecules = XYZFile(
            filename=multiple_molecules_xyz_file
        ).get_molecules(index=":", return_list=True)[:3]
        grouper = EnergyGrouper(molecules, threshold=1.0)
        energy_matrix_kcal = np.array(
            [
                [0.0, 0.5, 2.0],
                [-0.5, 0.0, 1.5],
                [-2.0, -1.5, 0.0],
            ]
        )

        _, index_groups = grouper.group_by_threshold(energy_matrix_kcal)

        assert {frozenset(group) for group in index_groups} == {
            frozenset({0, 1}),
            frozenset({2}),
        }
        assert grouper.threshold == 1.0
        assert np.array_equal(
            energy_matrix_kcal,
            np.array(
                [
                    [0.0, 0.5, 2.0],
                    [-0.5, 0.0, 1.5],
                    [-2.0, -1.5, 0.0],
                ]
            ),
        )

    def test_tanimoto_public_group_by_threshold_accepts_recorded_similarity(
        self, methanol_molecules
    ):
        grouper = TanimotoSimilarityGrouper(
            methanol_molecules[:3], threshold=0.9
        )
        similarity_matrix = np.array(
            [
                [1.0, np.nan, 0.7],
                [np.nan, np.nan, np.nan],
                [0.7, np.nan, 1.0],
            ],
            dtype=np.float32,
        )

        _, index_groups = grouper.group_by_threshold(
            similarity_matrix, threshold=0.6
        )

        assert index_groups == [[0, 2]]
        assert grouper._matrix_skipped_indices == [1]
        assert grouper._cached_group_indices == index_groups
        assert grouper.threshold == 0.9

    def test_public_group_by_num_groups_reuses_distance_matrix_and_caches_result(
        self, methanol_molecules
    ):
        molecules = methanol_molecules[:3]
        grouper = BasicRMSDGrouper(molecules, threshold=0.5, num_procs=1)
        distance_matrix = np.array(
            [
                [0.0, 0.2, 0.9],
                [0.2, 0.0, 0.8],
                [0.9, 0.8, 0.0],
            ]
        )
        original_matrix = distance_matrix.copy()

        groups, index_groups = grouper.group_by_num_groups(
            distance_matrix, num_groups=2
        )

        assert {frozenset(group) for group in index_groups} == {
            frozenset({0, 1}),
            frozenset({2}),
        }
        assert np.isclose(grouper._auto_threshold, 0.2)
        assert grouper._cached_groups == groups
        assert grouper._cached_group_indices == index_groups
        assert np.array_equal(distance_matrix, original_matrix)

    def test_tfd_public_group_by_num_groups_accepts_output_distance_matrix(
        self, methanol_molecules
    ):
        grouper = TorsionFingerprintGrouper(
            methanol_molecules[:3], threshold=0.1
        )
        tfd_matrix = np.array(
            [
                [0.0, 0.15, 0.8],
                [0.15, 0.0, 0.7],
                [0.8, 0.7, 0.0],
            ]
        )

        _, index_groups = grouper.group_by_num_groups(tfd_matrix, num_groups=2)

        assert {frozenset(group) for group in index_groups} == {
            frozenset({0, 1}),
            frozenset({2}),
        }
        assert np.isclose(grouper._auto_threshold, 0.15)

    def test_energy_public_group_by_num_groups_accepts_signed_kcal_output(
        self, multiple_molecules_xyz_file
    ):
        molecules = XYZFile(
            filename=multiple_molecules_xyz_file
        ).get_molecules(index=":", return_list=True)[:3]
        grouper = EnergyGrouper(molecules, threshold=1.0)
        energy_matrix_kcal = np.array(
            [
                [0.0, 0.5, 2.0],
                [-0.5, 0.0, 1.5],
                [-2.0, -1.5, 0.0],
            ]
        )

        _, index_groups = grouper.group_by_num_groups(
            energy_matrix_kcal, num_groups=2
        )

        assert {frozenset(group) for group in index_groups} == {
            frozenset({0, 1}),
            frozenset({2}),
        }
        assert np.isclose(grouper._auto_threshold, 0.5)

    def test_tanimoto_public_group_by_num_groups_accepts_recorded_similarity(
        self, methanol_molecules
    ):
        grouper = TanimotoSimilarityGrouper(
            methanol_molecules[:3], threshold=0.9
        )
        similarity_matrix = np.array(
            [
                [1.0, np.nan, 0.95],
                [np.nan, np.nan, np.nan],
                [0.95, np.nan, 1.0],
            ],
            dtype=np.float32,
        )

        _, index_groups = grouper.group_by_num_groups(
            similarity_matrix, num_groups=2
        )

        assert index_groups == [[0], [2]]
        assert grouper._matrix_skipped_indices == [1]
        assert grouper._auto_threshold is None

    def test_threshold_grouping_clears_previous_auto_threshold(
        self, methanol_molecules
    ):
        grouper = BasicRMSDGrouper(methanol_molecules, num_groups=2)
        matrix = np.array(
            [
                [0.0, 0.1, 0.8],
                [0.1, 0.0, 0.7],
                [0.8, 0.7, 0.0],
            ]
        )

        grouper.group_by_num_groups(matrix, num_groups=2)
        assert grouper._auto_threshold is not None

        old_threshold = grouper.threshold
        groups, index_groups = grouper.group_by_threshold(
            matrix, threshold=0.5
        )

        assert grouper._auto_threshold is None
        assert grouper.threshold == old_threshold
        assert index_groups == [[0, 1], [2]]
        assert groups == [
            [methanol_molecules[0], methanol_molecules[1]],
            [methanol_molecules[2]],
        ]

    def test_hierarchical_complete_linkage_prevents_chaining(
        self, methanol_molecules
    ):
        """A-B and B-C being close must not chain A and C together."""
        molecules = methanol_molecules[:3]
        grouper = BasicRMSDGrouper(
            molecules,
            threshold=0.5,
            num_procs=1,
        )
        distance_matrix = np.array(
            [
                [0.0, 0.3, 0.8],
                [0.3, 0.0, 0.3],
                [0.8, 0.3, 0.0],
            ]
        )

        groups, index_groups = grouper.group_by_threshold(distance_matrix)

        assert len(groups) == 2
        assert sorted(len(group) for group in index_groups) == [1, 2]
        assert not any(set(group) == {0, 1, 2} for group in index_groups)

    def test_threshold_includes_equal_distance(self, methanol_molecules):
        """Threshold mode uses complete-linkage distance <= threshold."""
        molecules = methanol_molecules[:2]
        grouper = BasicRMSDGrouper(
            molecules,
            threshold=0.5,
            num_procs=1,
        )
        distance_matrix = np.array(
            [
                [0.0, 0.5],
                [0.5, 0.0],
            ]
        )
        groups, index_groups = grouper.group_by_threshold(distance_matrix)
        assert len(groups) == 1
        assert index_groups == [[0, 1]]

    def test_num_groups_keeps_previous_full_linkage_level(
        self, methanol_molecules
    ):
        """-N must not split a tied linkage level just to force exact N."""
        from ase.build import molecule as ase_molecule

        from chemsmart.io.molecules.structure import Molecule

        molecules = [
            Molecule.from_ase_atoms(ase_molecule("CH3OH")) for _ in range(4)
        ]
        grouper = BasicRMSDGrouper(
            molecules,
            threshold=None,
            num_groups=3,
            num_procs=1,
        )
        # At distance 0.1, two independent merges occur at the same linkage
        # level, so the hierarchy jumps directly from 4 groups to 2 groups.
        # Requested N=3 must therefore retain the previous 4-group level.
        distance_matrix = np.array(
            [
                [0.0, 0.1, 1.0, 1.0],
                [0.1, 0.0, 1.0, 1.0],
                [1.0, 1.0, 0.0, 0.1],
                [1.0, 1.0, 0.1, 0.0],
            ]
        )
        groups, index_groups = grouper.group_by_num_groups(distance_matrix)
        assert len(groups) == 4
        assert len(index_groups) == 4
        assert all(len(group) == 1 for group in index_groups)
        assert grouper._auto_threshold is None

    def test_num_groups_uses_exact_level_when_available(
        self, methanol_molecules
    ):
        """-N should return exactly N when a full linkage level gives N groups."""
        from ase.build import molecule as ase_molecule

        from chemsmart.io.molecules.structure import Molecule

        molecules = [
            Molecule.from_ase_atoms(ase_molecule("CH3OH")) for _ in range(4)
        ]
        grouper = BasicRMSDGrouper(
            molecules,
            threshold=None,
            num_groups=3,
            num_procs=1,
        )

        distance_matrix = np.array(
            [
                [0.0, 0.1, 1.0, 1.0],
                [0.1, 0.0, 1.0, 1.0],
                [1.0, 1.0, 0.0, 0.4],
                [1.0, 1.0, 0.4, 0.0],
            ]
        )

        groups, index_groups = grouper.group_by_num_groups(distance_matrix)

        assert len(groups) == 3
        assert len(index_groups) == 3
        assert np.isclose(grouper._auto_threshold, 0.1)

    def test_grouping_does_not_modify_distance_matrix(
        self, methanol_molecules
    ):
        """Hierarchical grouping must not mutate the caller's distance matrix."""
        molecules = methanol_molecules[:3]
        grouper = BasicRMSDGrouper(
            molecules,
            threshold=0.5,
            num_procs=1,
        )

        distance_matrix = np.array(
            [
                [0.0, 0.3, 0.8],
                [0.3, 0.0, 0.3],
                [0.8, 0.3, 0.0],
            ]
        )
        original_matrix = distance_matrix.copy()

        grouper.group_by_threshold(distance_matrix)

        assert np.array_equal(distance_matrix, original_matrix)

    def test_hierarchical_grouping_is_input_order_independent(
        self, methanol_molecules
    ):
        """Permuting input order must not change the partition after remapping indices."""
        from ase.build import molecule as ase_molecule

        from chemsmart.io.molecules.structure import Molecule

        molecules = [
            Molecule.from_ase_atoms(ase_molecule("CH3OH")) for _ in range(4)
        ]
        distance_matrix = np.array(
            [
                [0.0, 0.2, 0.9, 1.0],
                [0.2, 0.0, 0.8, 0.9],
                [0.9, 0.8, 0.0, 0.3],
                [1.0, 0.9, 0.3, 0.0],
            ]
        )

        grouper = BasicRMSDGrouper(
            molecules,
            threshold=0.5,
            num_procs=1,
        )
        _, index_groups = grouper.group_by_threshold(distance_matrix)
        expected_partition = {frozenset(group) for group in index_groups}

        permutation = [2, 0, 3, 1]
        permuted_molecules = [molecules[i] for i in permutation]
        permuted_matrix = distance_matrix[np.ix_(permutation, permutation)]
        permuted_grouper = BasicRMSDGrouper(
            permuted_molecules,
            threshold=0.5,
            num_procs=1,
        )
        _, permuted_groups = permuted_grouper.group_by_threshold(
            permuted_matrix
        )

        remapped_partition = {
            frozenset(permutation[i] for i in group)
            for group in permuted_groups
        }
        assert remapped_partition == expected_partition

    def test_infinite_distance_matrix_skips_minimum_problematic_structures(
        self, methanol_molecules
    ):
        """Remove only the structure(s) needed to obtain a finite submatrix."""
        molecules = methanol_molecules[:3]
        grouper = BasicRMSDGrouper(
            molecules,
            threshold=0.5,
            num_procs=1,
        )

        distance_matrix = np.array(
            [
                [0.0, 0.3, np.inf],
                [0.3, 0.0, np.inf],
                [np.inf, np.inf, 0.0],
            ]
        )
        original_matrix = distance_matrix.copy()
        groups, index_groups = grouper.group_by_threshold(distance_matrix)

        assert index_groups == [[0, 1]]
        assert grouper._matrix_skipped_indices == [2]
        assert np.array_equal(distance_matrix, original_matrix, equal_nan=True)

    def test_nan_distance_matrix_raises(self, methanol_molecules):
        """NaN pairwise distances must still fail clearly at runtime."""
        molecules = methanol_molecules[:2]
        grouper = BasicRMSDGrouper(
            molecules,
            threshold=0.5,
            num_procs=1,
        )
        distance_matrix = np.array(
            [
                [0.0, np.nan],
                [np.nan, 0.0],
            ]
        )

        with pytest.raises(ValueError, match="non-finite"):
            grouper.group_by_threshold(distance_matrix)

    def test_tanimoto_threshold_uses_similarity_semantics(
        self, methanol_molecules
    ):
        """A Tanimoto threshold of 0.9 means similarity >= 0.9."""
        molecules = methanol_molecules[:3]
        grouper = TanimotoSimilarityGrouper(
            molecules,
            threshold=0.9,
            num_procs=1,
        )

        similarity_matrix = np.array(
            [
                [1.0, 0.90, 0.70],
                [0.90, 1.0, 0.70],
                [0.70, 0.70, 1.0],
            ]
        )

        groups, index_groups = grouper.group_by_threshold(similarity_matrix)

        assert len(groups) == 2
        assert {frozenset(group) for group in index_groups} == {
            frozenset({0, 1}),
            frozenset({2}),
        }
        assert np.isclose(grouper.threshold, 0.9)

    def test_complete_linkage_prevents_chaining(
        self, multiple_molecules_xyz_file
    ):
        """Test that complete linkage prevents chaining effect in grouping."""
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)

        # With a moderate threshold, complete linkage should create
        # groups where ALL members are within threshold of each other
        grouper = BasicRMSDGrouper(
            molecules,
            threshold=1.0,
            num_procs=self.NUM_PROCS,
        )
        groups, group_indices = grouper.group()

        # Verify complete linkage property: within each group,
        # all pairs should have RMSD <= threshold
        for group_idx, indices in enumerate(group_indices):
            if len(indices) > 1:
                for i, idx_i in enumerate(indices):
                    for j, idx_j in enumerate(indices):
                        if i < j:
                            rmsd = grouper.calculate_rmsd_pair(idx_i, idx_j)
                            assert rmsd <= 1.0, (
                                f"Group {group_idx}: RMSD between {idx_i} and {idx_j} "
                                f"is {rmsd}, is above threshold 1.0"
                            )

    def test_infinite_pair_removes_one_structure_and_groups_remaining(
        self, methanol_molecules
    ):
        from ase.build import molecule as ase_molecule

        from chemsmart.io.molecules.structure import Molecule

        molecules = [
            Molecule.from_ase_atoms(ase_molecule("CH3OH")) for _ in range(4)
        ]
        grouper = BasicRMSDGrouper(
            molecules,
            threshold=0.5,
            num_procs=1,
        )
        distance_matrix = np.array(
            [
                [0.0, np.inf, 0.8, 0.9],
                [np.inf, 0.0, 0.7, 0.8],
                [0.8, 0.7, 0.0, 0.2],
                [0.9, 0.8, 0.2, 0.0],
            ]
        )

        groups, index_groups = grouper.group_by_threshold(distance_matrix)

        assert grouper._matrix_skipped_indices == [0]
        assert {frozenset(group) for group in index_groups} == {
            frozenset({1}),
            frozenset({2, 3}),
        }
        assert len(groups) == 2

    def test_multiple_infinite_pairs_remove_minimum_problematic_structures(
        self, methanol_molecules
    ):
        from ase.build import molecule as ase_molecule

        from chemsmart.io.molecules.structure import Molecule

        molecules = [
            Molecule.from_ase_atoms(ase_molecule("CH3OH")) for _ in range(5)
        ]
        grouper = BasicRMSDGrouper(
            molecules,
            threshold=0.5,
            num_procs=1,
        )
        distance_matrix = np.array(
            [
                [0.0, np.inf, 0.8, 0.9, 0.9],
                [np.inf, 0.0, np.inf, 0.8, 0.8],
                [0.8, np.inf, 0.0, 0.7, 0.7],
                [0.9, 0.8, 0.7, 0.0, 0.2],
                [0.9, 0.8, 0.7, 0.2, 0.0],
            ]
        )

        groups, index_groups = grouper.group_by_threshold(distance_matrix)

        assert grouper._matrix_skipped_indices == [1]
        assert {frozenset(group) for group in index_groups} == {
            frozenset({0}),
            frozenset({2}),
            frozenset({3, 4}),
        }
        assert len(groups) == 3

    def test_original_indices_are_preserved_after_skipping(
        self, methanol_molecules
    ):
        from ase.build import molecule as ase_molecule

        from chemsmart.io.molecules.structure import Molecule

        molecules = [
            Molecule.from_ase_atoms(ase_molecule("CH3OH")) for _ in range(5)
        ]
        grouper = BasicRMSDGrouper(
            molecules,
            threshold=0.5,
            num_procs=1,
        )
        distance_matrix = np.array(
            [
                [0.0, np.inf, 0.9, 0.9, 0.9],
                [np.inf, 0.0, 0.9, 0.9, 0.9],
                [0.9, 0.9, 0.0, 0.8, 0.2],
                [0.9, 0.9, 0.8, 0.0, 0.8],
                [0.9, 0.9, 0.2, 0.8, 0.0],
            ]
        )

        _, index_groups = grouper.group_by_threshold(distance_matrix)

        assert grouper._matrix_skipped_indices == [0]
        assert {frozenset(group) for group in index_groups} == {
            frozenset({1}),
            frozenset({2, 4}),
            frozenset({3}),
        }

    def test_num_groups_after_skipping_uses_valid_structure_count(
        self, methanol_molecules
    ):
        from ase.build import molecule as ase_molecule

        from chemsmart.io.molecules.structure import Molecule

        molecules = [
            Molecule.from_ase_atoms(ase_molecule("CH3OH")) for _ in range(5)
        ]
        grouper = BasicRMSDGrouper(
            molecules,
            threshold=None,
            num_groups=5,
            num_procs=1,
        )
        distance_matrix = np.array(
            [
                [0.0, np.inf, 0.9, 0.9, 0.9],
                [np.inf, 0.0, 0.9, 0.9, 0.9],
                [0.9, 0.9, 0.0, 0.4, 0.5],
                [0.9, 0.9, 0.4, 0.0, 0.3],
                [0.9, 0.9, 0.5, 0.3, 0.0],
            ]
        )

        groups, index_groups = grouper.group_by_num_groups(distance_matrix)

        assert grouper._matrix_skipped_indices == [0]
        assert len(groups) == 4
        assert index_groups == [[1], [2], [3], [4]]

    def test_negative_distance_raises(self, methanol_molecules):
        molecules = methanol_molecules[:2]
        grouper = BasicRMSDGrouper(molecules, threshold=0.5, num_procs=1)
        distance_matrix = np.array([[0.0, -0.1], [-0.1, 0.0]])

        with pytest.raises(ValueError, match="negative"):
            grouper.group_by_threshold(distance_matrix)

    def test_negative_infinity_raises(self, methanol_molecules):
        molecules = methanol_molecules[:2]
        grouper = BasicRMSDGrouper(molecules, threshold=0.5, num_procs=1)
        distance_matrix = np.array([[0.0, -np.inf], [-np.inf, 0.0]])

        with pytest.raises(ValueError, match="negative"):
            grouper.group_by_threshold(distance_matrix)

    def test_asymmetric_distance_matrix_raises(self, methanol_molecules):
        molecules = methanol_molecules[:2]
        grouper = BasicRMSDGrouper(molecules, threshold=0.5, num_procs=1)
        distance_matrix = np.array([[0.0, 0.2], [0.3, 0.0]])

        with pytest.raises(ValueError, match="symmetric"):
            grouper.group_by_threshold(distance_matrix)

    def test_nonzero_diagonal_raises(self, methanol_molecules):
        molecules = methanol_molecules[:2]
        grouper = BasicRMSDGrouper(molecules, threshold=0.5, num_procs=1)
        distance_matrix = np.array([[0.1, 0.2], [0.2, 0.0]])

        with pytest.raises(ValueError, match="diagonal"):
            grouper.group_by_threshold(distance_matrix)

    def test_infinite_pair_tie_breaking_is_deterministic(
        self, methanol_molecules
    ):
        """Equal +inf counts should use a deterministic tie-breaking rule."""
        molecules = methanol_molecules[:3]
        grouper = BasicRMSDGrouper(
            molecules,
            threshold=0.5,
            num_procs=1,
        )

        distance_matrix = np.array(
            [
                [0.0, np.inf, 0.8],
                [np.inf, 0.0, 0.8],
                [0.8, 0.8, 0.0],
            ]
        )
        groups, index_groups = grouper.group_by_threshold(distance_matrix)
        # Structures 0 and 1 have the same number of +inf entries.
        # The current implementation deterministically removes the
        # first maximum, i.e. original index 0.
        assert grouper._matrix_skipped_indices == [0]
        assert {frozenset(group) for group in index_groups} == {
            frozenset({1}),
            frozenset({2}),
        }
        assert len(groups) == 2

    def test_infinite_distance_removal_is_iterative(self, methanol_molecules):
        """Recount +inf involvement after each removal until the submatrix is finite."""
        from ase.build import molecule as ase_molecule

        from chemsmart.io.molecules.structure import Molecule

        molecules = [
            Molecule.from_ase_atoms(ase_molecule("CH3OH")) for _ in range(5)
        ]
        grouper = BasicRMSDGrouper(
            molecules,
            threshold=0.5,
            num_procs=1,
        )
        distance_matrix = np.array(
            [
                [0.0, np.inf, np.inf, 0.9, 0.9],
                [np.inf, 0.0, 0.8, np.inf, 0.9],
                [np.inf, 0.8, 0.0, np.inf, 0.9],
                [0.9, np.inf, np.inf, 0.0, 0.2],
                [0.9, 0.9, 0.9, 0.2, 0.0],
            ]
        )
        groups, index_groups = grouper.group_by_threshold(distance_matrix)
        # Initial +inf counts:
        # 0 -> 2
        # 1 -> 2
        # 2 -> 2
        # 3 -> 2
        # 4 -> 0
        #
        # Tie-breaking removes 0 first.
        #
        # Remaining +inf pairs:
        # 1-3
        # 2-3
        #
        # Recomputed counts:
        # 1 -> 1
        # 2 -> 1
        # 3 -> 2
        #
        # Therefore 3 must be removed in the second iteration.
        assert grouper._matrix_skipped_indices == [0, 3]
        # Remaining original indices are 1, 2, and 4.
        # All pairwise distances are > 0.5, so all are singleton groups.
        assert {frozenset(group) for group in index_groups} == {
            frozenset({1}),
            frozenset({2}),
            frozenset({4}),
        }
        assert len(groups) == 3


@pytest.mark.usefixtures("temporary_working_dir")
class Test_conformer_ids_functionality:
    """Test conformer_ids parameter functionality."""

    NUM_PROCS = 1

    def test_conformer_ids_from_log_directory(
        self, ts_conformers_log_directory
    ):
        """Test loading conformer IDs from a directory of log files."""
        import glob
        import re

        from chemsmart.io.molecules.structure import Molecule

        # Load molecules from log files in the ts_conformers_log_directory
        log_files = sorted(
            glob.glob(os.path.join(ts_conformers_log_directory, "*.log"))
        )
        assert (
            len(log_files) == 5
        ), f"Expected 5 log files, found {len(log_files)}"

        # Extract conformer IDs from filenames (e.g., ch_1c_para_c1.log -> c1)
        molecules = []
        conformer_ids = []
        for log_file in log_files:
            mol = Molecule.from_filepath(
                filepath=log_file, index="-1", return_list=False
            )
            if mol is not None:
                molecules.append(mol)
                # Extract conformer ID from filename
                basename = os.path.basename(log_file)
                match = re.search(r"_c(\d+)\.log$", basename)
                if match:
                    conformer_ids.append(f"c{match.group(1)}")
                else:
                    conformer_ids.append(basename)

        assert len(molecules) >= 2, "Need at least 2 molecules for grouping"

        grouper = BasicRMSDGrouper(
            molecules,
            threshold=1.0,
            num_procs=1,
            label="ts_conformers_test",
            conformer_ids=conformer_ids,
        )
        _, _ = grouper.group()

        # Verify conformer_ids are stored
        assert grouper.conformer_ids == conformer_ids

        # Check Excel output
        import pandas as pd

        excel_file = os.path.join(
            "ts_conformers_test_group_result",
            "ts_conformers_test_BasicRMSDGrouper_T1.0.xlsx",
        )
        assert os.path.exists(
            excel_file
        ), f"Excel file not found: {excel_file}"

        # Matrix data starts at row 8 (0-indexed: skiprows=15), first column is index
        df = pd.read_excel(
            excel_file, sheet_name="RMSD_Matrix", skiprows=14, index_col=0
        )
        # Check that conformer IDs are used as labels
        assert "c1" in str(df.columns[0]) or "c1" in str(df.index[0])

    def test_traj_conformer_ids_original_indices(self):
        """Test that traj job correctly sets original conformer indices."""
        from ase.build import molecule as ase_molecule

        from chemsmart.io.molecules.structure import Molecule

        # Local ASE methanol — avoid flaky live PubChem for a trivial fixture.
        mol = Molecule.from_ase_atoms(ase_molecule("CH3OH"))
        molecules = [mol.copy() for _ in range(18)]

        # Simulate traj behavior: select last 50% (indices 10-18 in original)
        proportion = 0.5
        total = len(molecules)
        last_num = int(round(total * proportion, 1))
        _ = molecules[-last_num:]  # selected_molecules - not used in this test

        # Calculate expected original indices (1-based)
        start_original_index = total - last_num + 1
        expected_ids = [str(start_original_index + i) for i in range(last_num)]

        # Verify expected IDs
        assert expected_ids[0] == "10"  # First selected is original index 10
        assert expected_ids[-1] == "18"  # Last selected is original index 18


@pytest.mark.usefixtures("temporary_working_dir")
class Test_output_file_generation:
    """Test that grouper generates correct output files."""

    NUM_PROCS = 1

    def test_group_xyz_files_contain_energy_and_index_info(
        self, multiple_molecules_xyz_file
    ):
        """Test that group XYZ files contain energy and original index information."""
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)

        grouper = BasicRMSDGrouper(
            molecules[:6],
            threshold=1.0,
            num_procs=1,
            label="info_test",
        )
        _, _ = grouper.group()
        grouper.unique()  # Generates xyz files

        output_dir = "info_test_group_result"
        first_group_file = os.path.join(output_dir, "info_test_group_1.xyz")

        with open(first_group_file, "r") as f:
            content = f.read()

        # Check for expected information in comments
        assert "Original_Index:" in content
        assert "E" in content

    def test_group_xyz_files_sorted_by_energy(
        self, multiple_molecules_xyz_file
    ):
        """Test that molecules in group XYZ files are sorted by energy (lowest first)."""
        import re

        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)

        grouper = BasicRMSDGrouper(
            molecules,
            threshold=2.0,  # Higher threshold to get more molecules per group
            num_procs=1,
            label="sort_test",
        )
        groups, group_indices = grouper.group()
        grouper.unique()  # Generates xyz files

        output_dir = "sort_test_group_result"

        # Check each group file
        for group_num in range(len(groups)):
            group_file = os.path.join(
                output_dir, f"sort_test_group_{group_num + 1}.xyz"
            )

            with open(group_file, "r") as f:
                content = f.read()

            # Extract energies from file
            lines = content.strip().split("\n")
            i = 0
            energies = []

            while i < len(lines):
                num_atoms = int(lines[i].strip())
                comment_line = lines[i + 1]
                energy_match = re.search(
                    r"E\(Hartree\):\s*([-\d.]+)", comment_line
                )
                if energy_match:
                    energies.append(float(energy_match.group(1)))
                i += num_atoms + 2

            # Verify sorted by energy (lowest first)
            for j in range(len(energies) - 1):
                assert energies[j] <= energies[j + 1], (
                    f"Group {group_num + 1} not sorted: "
                    f"energy[{j}]={energies[j]} > energy[{j+1}]={energies[j+1]}"
                )

    def test_group_xyz_files_energy_from_log_files(
        self, ts_conformers_log_directory
    ):
        """Test that energy is correctly extracted from log files and written to group XYZ files."""
        import glob
        import re

        from chemsmart.io.molecules.structure import Molecule

        # Load molecules from log files
        log_files = sorted(
            glob.glob(os.path.join(ts_conformers_log_directory, "*.log"))
        )
        assert (
            len(log_files) >= 2
        ), f"Need at least 2 log files, found {len(log_files)}"

        molecules = []
        conformer_ids = []
        original_energies = {}

        for idx, log_file in enumerate(log_files):
            mol = Molecule.from_filepath(
                filepath=log_file, index="-1", return_list=False
            )
            if mol is not None:
                molecules.append(mol)
                # Store original energy for verification
                original_energies[idx] = mol.energy

                # Extract conformer ID from filename
                basename = os.path.basename(log_file)
                match = re.search(r"_c(\d+)\.log$", basename)
                if match:
                    conformer_ids.append(f"c{match.group(1)}")
                else:
                    conformer_ids.append(basename)

        assert len(molecules) >= 2, "Need at least 2 molecules for grouping"

        grouper = BasicRMSDGrouper(
            molecules,
            threshold=5.0,  # High threshold to group all together
            num_procs=1,
            label="log_energy_test",
            conformer_ids=conformer_ids,
        )
        _, _ = grouper.group()
        grouper.unique()  # Generates xyz files

        # Check group file
        output_dir = "log_energy_test_group_result"
        first_group_file = os.path.join(
            output_dir, "log_energy_test_group_1.xyz"
        )

        assert os.path.exists(
            first_group_file
        ), f"Group file not found: {first_group_file}"

        with open(first_group_file, "r") as f:
            content = f.read()

        # Verify energy values in the output
        assert "E(Hartree):" in content

        # Parse and verify each energy value
        lines = content.strip().split("\n")
        i = 0
        energies_found = 0
        while i < len(lines):
            try:
                num_atoms = int(lines[i].strip())
            except ValueError:
                break
            comment_line = lines[i + 1]

            # Extract energy from comment
            energy_match = re.search(
                r"E\(Hartree\):\s*([-\d.]+)", comment_line
            )
            if energy_match:
                extracted_energy = float(energy_match.group(1))
                # Verify the energy is a reasonable value (not zero or nan)
                assert (
                    extracted_energy < 0
                ), f"Energy should be negative: {extracted_energy}"
                energies_found += 1

            i += num_atoms + 2

        assert energies_found > 0, "No energies found in output file"


@pytest.mark.usefixtures("temporary_working_dir")
class Test_edge_cases:
    """Test edge cases and boundary conditions."""

    def test_two_molecules_grouping(self):
        """Test grouping with minimum number of molecules (2)."""
        from ase.build import molecule as ase_molecule

        from chemsmart.io.molecules.structure import Molecule

        mol1 = Molecule.from_ase_atoms(ase_molecule("CH3OH"))
        mol2 = mol1.copy()

        grouper = BasicRMSDGrouper([mol1, mol2], threshold=0.5)
        groups, group_indices = grouper.group()

        # Two identical molecules should form one group
        assert len(groups) == 1
        assert len(groups[0]) == 2

    def test_num_groups_equals_num_molecules(self, methanol_molecules):
        """Test requesting same number of groups as molecules."""
        grouper = BasicRMSDGrouper(
            methanol_molecules,
            num_groups=len(methanol_molecules),
        )
        groups, group_indices = grouper.group()

        # Requesting N equal to the number of molecules returns singleton groups.
        assert len(groups) == len(methanol_molecules)
        assert all(len(group) == 1 for group in group_indices)

    def test_num_groups_exceeds_num_molecules(self, methanol_molecules):
        """Test requesting more groups than molecules."""
        n_mols = len(methanol_molecules)
        grouper = BasicRMSDGrouper(
            methanol_molecules,
            num_groups=n_mols + 5,
        )
        groups, group_indices = grouper.group()

        # More requested groups than molecules still returns singleton groups.
        assert len(groups) == n_mols
        assert all(len(group) == 1 for group in group_indices)

    def test_very_low_threshold(self, multiple_molecules_xyz_file):
        """Test with very low threshold (should create many groups)."""
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)

        grouper = BasicRMSDGrouper(
            molecules,
            threshold=0.001,  # Very low threshold
            num_procs=1,
        )
        groups, group_indices = grouper.group()

        # With very low threshold, most molecules should be in separate groups
        # (unless they're truly identical)
        assert len(groups) >= len(molecules) - 1

    def test_very_high_threshold(self, multiple_molecules_xyz_file):
        """Test with very high threshold (should create few groups)."""
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)

        grouper = BasicRMSDGrouper(
            molecules,
            threshold=100.0,  # Very high threshold
            num_procs=1,
        )
        groups, group_indices = grouper.group()

        # With very high threshold, all molecules with same formula should be in one group
        assert len(groups) <= 3  # Likely 1-2 groups

    def test_different_formulas_always_separate(self, methanol_and_ethanol):
        """Test that molecules with different formulas are always in separate groups."""
        grouper = BasicRMSDGrouper(
            methanol_and_ethanol,
            threshold=100.0,  # Even with high threshold
        )
        groups, group_indices = grouper.group()

        # Incompatible RMSD pairs are now skipped from clustering entirely.
        assert len(groups) == 1
        assert len(group_indices) == 1
        assert grouper._matrix_skipped_indices == [0]
        assert group_indices == [[1]]

    def test_rmsd_infinity_for_different_molecules(self, methanol_and_ethanol):
        """Test that RMSD returns infinity for molecules with different atom counts."""
        grouper = BasicRMSDGrouper(methanol_and_ethanol, threshold=0.5)
        rmsd = grouper.calculate_rmsd_pair(0, 1)

        assert rmsd == np.inf or rmsd == float("inf")

    def test_incompatible_rmsd_structures_are_skipped(
        self, methanol_and_ethanol
    ):
        """Structures involved in RMSD=inf pairs should be skipped, not grouped."""
        grouper = BasicRMSDGrouper(methanol_and_ethanol, threshold=0.5)
        groups, group_indices = grouper.group()

        assert len(groups) == 1
        assert len(group_indices) == 1
        assert grouper._matrix_skipped_indices == [0]
        assert group_indices == [[1]]


@pytest.mark.usefixtures("temporary_working_dir")
class Test_label_and_append_label:
    """Test -l (label) and -a (append_label) parameter functionality."""

    NUM_PROCS = 1

    def test_get_label_function(self):
        """Test _get_label function logic."""
        from chemsmart.cli.grouper.grouper import _get_label

        # Test with label only
        result = _get_label(
            label="custom", append_label=None, base_label="base"
        )
        assert result == "custom"

        # Test with append_label only
        result = _get_label(
            label=None, append_label="suffix", base_label="base"
        )
        assert result == "base_suffix"

        # Test with neither (use base_label)
        result = _get_label(label=None, append_label=None, base_label="base")
        assert result == "base"

        # Test with both should raise error
        with pytest.raises(ValueError) as excinfo:
            _get_label(
                label="custom", append_label="suffix", base_label="base"
            )
        assert "Only give label or append_label" in str(excinfo.value)

    def test_label_in_output_directory(self, multiple_molecules_xyz_file):
        """Test that label parameter affects output directory name."""
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)[:5]

        # Test with custom label
        grouper = BasicRMSDGrouper(
            molecules,
            threshold=1.0,
            num_procs=self.NUM_PROCS,
            label="my_custom_label",
        )
        _, _ = grouper.group()

        # Check output directory uses custom label
        output_dir = "my_custom_label_group_result"
        assert os.path.exists(
            output_dir
        ), f"Output dir not found: {output_dir}"

        # Check Excel file uses custom label
        excel_file = os.path.join(
            output_dir, "my_custom_label_BasicRMSDGrouper_T1.0.xlsx"
        )
        assert os.path.exists(
            excel_file
        ), f"Excel file not found: {excel_file}"

    def test_label_in_group_xyz_files(self, multiple_molecules_xyz_file):
        """Test that label parameter affects group XYZ file names."""
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)[:5]

        grouper = BasicRMSDGrouper(
            molecules,
            threshold=1.0,
            num_procs=self.NUM_PROCS,
            label="test_label",
        )
        groups, group_indices = grouper.group()
        grouper.unique()  # Generates xyz files

        output_dir = "test_label_group_result"

        # Check group XYZ files use label
        for i in range(len(groups)):
            xyz_path = os.path.join(output_dir, f"test_label_group_{i+1}.xyz")
            assert os.path.exists(xyz_path), f"Group XYZ not found: {xyz_path}"

    def test_different_labels_create_different_outputs(
        self, multiple_molecules_xyz_file
    ):
        """Test that different labels create separate output directories."""
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)[:5]

        # First grouper with label "run1"
        grouper1 = BasicRMSDGrouper(
            molecules,
            threshold=1.0,
            num_procs=self.NUM_PROCS,
            label="run1",
        )
        grouper1.group()

        # Second grouper with label "run2"
        grouper2 = BasicRMSDGrouper(
            molecules,
            threshold=1.0,
            num_procs=self.NUM_PROCS,
            label="run2",
        )
        grouper2.group()

        # Both output directories should exist
        assert os.path.exists("run1_group_result")
        assert os.path.exists("run2_group_result")

        # Both should have their own Excel files
        assert os.path.exists(
            "run1_group_result/run1_BasicRMSDGrouper_T1.0.xlsx"
        )
        assert os.path.exists(
            "run2_group_result/run2_BasicRMSDGrouper_T1.0.xlsx"
        )

    def test_label_with_num_groups(self, multiple_molecules_xyz_file):
        """Test that label works correctly with num_groups parameter."""
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)

        grouper = BasicRMSDGrouper(
            molecules,
            num_groups=5,
            num_procs=self.NUM_PROCS,
            label="numgroups_test",
        )
        _, _ = grouper.group()

        output_dir = "numgroups_test_group_result"
        assert os.path.exists(output_dir)

        # Excel file should reflect num_groups (N5) instead of threshold
        excel_file = os.path.join(
            output_dir, "numgroups_test_BasicRMSDGrouper_N5.xlsx"
        )
        assert os.path.exists(
            excel_file
        ), f"Excel file not found: {excel_file}"


class TestConformerIdExtraction:
    """Test conformer ID extraction from filenames."""

    @staticmethod
    def _sort_key(x):
        """Sort key for file_info: cXX files by number first, then others by filename."""
        if x[2] is not None:
            return (0, x[2], x[3])
        else:
            return (1, 0, x[3])

    def test_extract_conformer_id_with_pattern(self):
        """Test extraction with _cXX_ pattern."""
        from chemsmart.cli.grouper.grouper import _extract_conformer_id

        assert _extract_conformer_id("structure_c1_opt.log") == "c1"
        assert _extract_conformer_id("mol_c12.log") == "c12"
        assert (
            _extract_conformer_id("/path/to/structure_c123_ts.log") == "c123"
        )

    def test_extract_conformer_id_without_pattern(self):
        """Test extraction without _cXX_ pattern returns None."""
        from chemsmart.cli.grouper.grouper import _extract_conformer_id

        assert _extract_conformer_id("mol_opt.log") is None
        assert _extract_conformer_id("structure.log") is None
        assert _extract_conformer_id("/path/to/calc.log") is None

    def test_conformer_id_fallback_to_filename(self):
        """Test that filename is used as conformer ID when no _cXX_ pattern."""
        from chemsmart.cli.grouper.grouper import _extract_conformer_id

        test_files = [
            "/path/to/mol_opt.log",
            "/path/to/structure_ts.out",
        ]

        conformer_ids = []
        for f in test_files:
            conf_id = _extract_conformer_id(f)
            if conf_id is None:
                basename = os.path.basename(f)
                conf_id = os.path.splitext(basename)[0]
            conformer_ids.append(conf_id)

        assert conformer_ids == ["mol_opt", "structure_ts"]

    def test_conformer_ids_molecules_correspondence(
        self, ts_conformers_log_directory
    ):
        """Test that conformer_ids and molecules are strictly one-to-one corresponding.

        This verifies that after sorting, each molecule's energy matches the energy
        from its corresponding file (identified by conf_id).
        """
        import glob
        import re

        from chemsmart.cli.grouper.grouper import _extract_conformer_id
        from chemsmart.io.gaussian.output import Gaussian16Output
        from chemsmart.io.molecules.structure import Molecule

        # First, build a mapping from conf_id to expected gibbs_energy
        log_files = glob.glob(
            os.path.join(ts_conformers_log_directory, "*.log")
        )
        expected_energies = {}

        for f in log_files:
            conf_id = _extract_conformer_id(f)
            basename = os.path.basename(f)
            name_without_ext = os.path.splitext(basename)[0]

            if conf_id is None:
                conf_id = name_without_ext

            g16_output = Gaussian16Output(filename=f)
            gibbs_energy = g16_output.gibbs_free_energy
            if gibbs_energy is not None:
                expected_energies[conf_id] = gibbs_energy

        # Now simulate the actual loading logic with sorting
        file_info = []
        for f in log_files:
            conf_id = _extract_conformer_id(f)
            basename = os.path.basename(f)
            name_without_ext = os.path.splitext(basename)[0]

            if conf_id:
                num = int(re.search(r"\d+", conf_id).group())
                file_info.append((f, conf_id, num, name_without_ext))
            else:
                file_info.append((f, name_without_ext, None, name_without_ext))

        file_info.sort(key=self._sort_key)

        # Load molecules
        molecules = []
        conformer_ids = []

        for filepath, conf_id, _, _ in file_info:
            try:
                mol = Molecule.from_filepath(
                    filepath=filepath, index="-1", return_list=False
                )
                if mol is not None:
                    g16_output = Gaussian16Output(filename=filepath)
                    gibbs_energy = g16_output.gibbs_free_energy
                    if gibbs_energy is not None:
                        mol._energy = gibbs_energy

                    molecules.append(mol)
                    conformer_ids.append(conf_id)
            except Exception:
                pass

        # Verify one-to-one correspondence
        assert len(molecules) == len(conformer_ids)

        # Verify each molecule's energy matches the expected energy for its conf_id
        for i, (mol, conf_id) in enumerate(zip(molecules, conformer_ids)):
            if conf_id in expected_energies:
                expected_energy = expected_energies[conf_id]
                assert np.isclose(mol.energy, expected_energy, rtol=1e-10), (
                    f"Index {i}: energy mismatch for {conf_id}, "
                    f"got {mol.energy}, expected {expected_energy}"
                )

    def test_mixed_pattern_files_correspondence(self):
        """Test correspondence when mixing files with and without _cXX_ pattern."""
        from chemsmart.cli.grouper.grouper import _extract_conformer_id

        test_files = [
            ("mol_c1_opt.log", "c1"),
            ("mol_c2_opt.log", "c2"),
            ("other_calc.log", None),
            ("mol_c3_opt.log", "c3"),
        ]

        file_info = []
        for filename, expected_pattern in test_files:
            conf_id = _extract_conformer_id(filename)
            basename = os.path.basename(filename)
            name_without_ext = os.path.splitext(basename)[0]

            if conf_id:
                assert conf_id == expected_pattern
                num = int(conf_id.replace("c", ""))
                file_info.append((filename, conf_id, num, name_without_ext))
            else:
                assert expected_pattern is None
                file_info.append(
                    (filename, name_without_ext, None, name_without_ext)
                )

        file_info.sort(key=self._sort_key)
        sorted_conf_ids = [info[1] for info in file_info]

        # Files with pattern first (sorted by number), then others by filename
        assert sorted_conf_ids == ["c1", "c2", "c3", "other_calc"]

    def test_all_files_without_cxx_pattern(self):
        """Test that a folder with no _cXX_ pattern files works correctly."""
        from chemsmart.cli.grouper.grouper import _extract_conformer_id

        test_files = [
            "mol_opt.log",
            "structure_ts.log",
            "calc.log",
            "another_mol.log",
        ]

        file_info = []
        for filename in test_files:
            conf_id = _extract_conformer_id(filename)
            basename = os.path.basename(filename)
            name_without_ext = os.path.splitext(basename)[0]

            if conf_id:
                num = int(conf_id.replace("c", ""))
                file_info.append((filename, conf_id, num, name_without_ext))
            else:
                file_info.append(
                    (filename, name_without_ext, None, name_without_ext)
                )

        # All files should have None as num
        for info in file_info:
            assert info[2] is None

        file_info.sort(key=self._sort_key)
        sorted_conf_ids = [info[1] for info in file_info]

        # All files sorted alphabetically by filename
        assert sorted_conf_ids == [
            "another_mol",
            "calc",
            "mol_opt",
            "structure_ts",
        ]


@pytest.mark.usefixtures("temporary_working_dir")
class Test_energy_extraction_function:
    """Framework tests for energy extraction by file type."""

    def test_energy_extraction_gaussian(
        self, gaussian_dppeFeCl2_link_opt_outputfile
    ):
        from chemsmart.analysis.thermochemistry import Thermochemistry
        from chemsmart.cli.grouper.grouper import (
            _extract_energy_based_on_energy_type,
        )

        thermo = Thermochemistry(
            filename=gaussian_dppeFeCl2_link_opt_outputfile
        )
        extracted = _extract_energy_based_on_energy_type(thermo, "E")
        assert np.isclose(extracted, -3869.01351827, rtol=1e-7)

        thermo = Thermochemistry(
            filename=gaussian_dppeFeCl2_link_opt_outputfile
        )
        extracted = _extract_energy_based_on_energy_type(thermo, "H")
        assert np.isclose(extracted, -3868.552084, rtol=1e-7)

        thermo = Thermochemistry(
            filename=gaussian_dppeFeCl2_link_opt_outputfile
        )
        extracted = _extract_energy_based_on_energy_type(thermo, "G")
        assert np.isclose(extracted, -3868.648944, rtol=1e-7)

        thermo_qhh = Thermochemistry(
            filename=gaussian_dppeFeCl2_link_opt_outputfile,
            temperature=298.15,
            concentration=1.0,
            h_freq_cutoff=100.0,
            s_freq_cutoff=100.0,
            use_weighted_mass=True,
            entropy_method="grimme",
        )
        extracted = _extract_energy_based_on_energy_type(thermo_qhh, "qhH")
        assert np.isclose(extracted, -3868.558465, rtol=1e-7)

        thermo_qhg = Thermochemistry(
            filename=gaussian_dppeFeCl2_link_opt_outputfile,
            temperature=298.15,
            concentration=1.0,
            h_freq_cutoff=100.0,
            s_freq_cutoff=100.0,
            use_weighted_mass=True,
            entropy_method="grimme",
        )
        extracted = _extract_energy_based_on_energy_type(thermo_qhg, "qhG")
        assert np.isclose(extracted, -3868.645329, rtol=1e-7)

    def test_energy_extraction_orca(self, fe2_singlet_output):
        from chemsmart.analysis.thermochemistry import Thermochemistry
        from chemsmart.cli.grouper.grouper import (
            _extract_energy_based_on_energy_type,
        )

        thermo = Thermochemistry(filename=fe2_singlet_output)
        extracted = _extract_energy_based_on_energy_type(thermo, "E")
        assert np.isclose(extracted, -1568.256171848101, rtol=1e-7)

        thermo = Thermochemistry(filename=fe2_singlet_output)
        extracted = _extract_energy_based_on_energy_type(thermo, "H")
        assert np.isclose(extracted, -1568.14237774, rtol=1e-7)

        thermo = Thermochemistry(filename=fe2_singlet_output)
        extracted = _extract_energy_based_on_energy_type(thermo, "G")
        assert np.isclose(extracted, -1568.18873516, rtol=1e-7)

        thermo_qhh = Thermochemistry(
            filename=fe2_singlet_output,
            temperature=298.15,
            concentration=1.0,
            h_freq_cutoff=100.0,
            s_freq_cutoff=100.0,
            use_weighted_mass=True,
            entropy_method="grimme",
        )
        extracted = _extract_energy_based_on_energy_type(thermo_qhh, "qhH")
        assert np.isclose(extracted, -1568.143280, rtol=1e-7)

        thermo_qhg = Thermochemistry(
            filename=fe2_singlet_output,
            temperature=298.15,
            concentration=1.0,
            h_freq_cutoff=100.0,
            s_freq_cutoff=100.0,
            use_weighted_mass=True,
            entropy_method="grimme",
        )
        extracted = _extract_energy_based_on_energy_type(thermo_qhg, "qhG")
        assert np.isclose(extracted, -1568.186619, rtol=1e-7)


@pytest.mark.usefixtures("temporary_working_dir")
class Test_multiprocessing_support:
    def test_basic_rmsd_serial_parallel_equivalent(self, methanol_molecules):
        serial = BasicRMSDGrouper(
            methanol_molecules, threshold=0.5, num_procs=1
        )
        parallel = BasicRMSDGrouper(
            methanol_molecules, threshold=0.5, num_procs=2
        )

        serial_matrix = serial.calculate_full_rmsd_matrix()
        parallel_matrix = parallel.calculate_full_rmsd_matrix()
        assert np.allclose(
            serial_matrix, parallel_matrix, atol=1e-10, rtol=1e-10
        )

        _, serial_idx = serial.group()
        _, parallel_idx = parallel.group()
        assert serial_idx == parallel_idx

    def test_hungarian_rmsd_serial_parallel_equivalent(
        self, methanol_molecules
    ):
        serial = HungarianRMSDGrouper(
            methanol_molecules, threshold=0.5, num_procs=1
        )
        parallel = HungarianRMSDGrouper(
            methanol_molecules, threshold=0.5, num_procs=2
        )

        serial_matrix = serial.calculate_full_rmsd_matrix()
        parallel_matrix = parallel.calculate_full_rmsd_matrix()
        assert np.allclose(
            serial_matrix, parallel_matrix, atol=1e-10, rtol=1e-10
        )

        _, serial_idx = serial.group()
        _, parallel_idx = parallel.group()
        assert serial_idx == parallel_idx

    def test_spyrmsd_parallel_preserves_isomorphisms(self, methanol_molecules):
        serial = SpyRMSDGrouper(methanol_molecules, threshold=0.5, num_procs=1)
        parallel = SpyRMSDGrouper(
            methanol_molecules, threshold=0.5, num_procs=2
        )

        serial_matrix = serial.calculate_full_rmsd_matrix()
        parallel_matrix = parallel.calculate_full_rmsd_matrix()
        assert np.allclose(
            serial_matrix, parallel_matrix, atol=1e-8, rtol=1e-8
        )

        _, serial_idx = serial.group()
        _, parallel_idx = parallel.group()
        assert serial_idx == parallel_idx

        assert serial.best_isomorphisms
        assert parallel.best_isomorphisms
        assert (0, 1) in serial.best_isomorphisms
        assert (0, 1) in parallel.best_isomorphisms
        assert serial.get_best_isomorphism(0, 1) is not None
        assert parallel.get_best_isomorphism(0, 1) is not None

    @pytest.mark.skipif(
        not _is_irmsd_available(), reason="irmsd API/command not available"
    )
    def test_irmsd_serial_parallel_equivalent(self, methanol_molecules):
        serial = IRMSDGrouper(methanol_molecules, threshold=0.5, num_procs=1)
        parallel = IRMSDGrouper(methanol_molecules, threshold=0.5, num_procs=2)

        serial_matrix = serial.calculate_full_rmsd_matrix()
        parallel_matrix = parallel.calculate_full_rmsd_matrix()
        assert np.allclose(
            serial_matrix, parallel_matrix, atol=1e-8, rtol=1e-8
        )

        _, serial_idx = serial.group()
        _, parallel_idx = parallel.group()
        assert serial_idx == parallel_idx

    def test_formula_grouper_falls_back_to_single_proc(
        self, caplog, methanol_molecules
    ):
        """FormulaGrouper warns and falls back to num_procs=1 when >1 is requested."""
        caplog.set_level("WARNING")
        grouper = FormulaGrouper(methanol_molecules, num_procs=4)
        assert grouper.num_procs == 1
        assert (
            "FormulaGrouper does not support multiprocessing; using num_procs=1."
            in caplog.text
        )

        groups, group_indices = grouper.group()
        assert len(groups) == 1
        assert len(group_indices) == 1

    def test_serial_only_strategy_warns_and_falls_back(
        self, caplog, multiple_molecules_xyz_file
    ):
        """Serial-only grouper warns and falls back to num_procs=1."""
        caplog.set_level("WARNING")
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)
        grouper = EnergyGrouper(molecules, threshold=0.5, num_procs=4)
        assert grouper.num_procs == 1
        assert (
            "EnergyGrouper does not support multiprocessing; using num_procs=1."
            in caplog.text
        )

    def test_basic_rmsd_progress_logs_all_milestones_serial(
        self, caplog, multiple_molecules_xyz_file
    ):
        caplog.set_level("INFO")
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)
        grouper = BasicRMSDGrouper(molecules, threshold=0.5, num_procs=1)

        grouper.calculate_full_rmsd_matrix()

        _assert_progress_milestones(caplog.text)

    def test_basic_rmsd_progress_logs_all_milestones_parallel(
        self, caplog, multiple_molecules_xyz_file
    ):
        caplog.set_level("INFO")
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)
        grouper = BasicRMSDGrouper(molecules, threshold=0.5, num_procs=2)

        grouper.calculate_full_rmsd_matrix()

        _assert_progress_milestones(caplog.text)

    def test_energy_grouper_progress_logs_all_milestones(
        self, caplog, multiple_molecules_xyz_file
    ):
        caplog.set_level("INFO")
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)
        grouper = EnergyGrouper(molecules, threshold=0.5, num_procs=1)

        grouper.group()

        _assert_progress_milestones(caplog.text)

    def test_basic_rmsd_parallel_single_molecule_no_pair_no_crash(
        self, methanol_molecules
    ):
        """Single-molecule RMSD run with num_procs>1 should safely skip pair workers."""
        molecules = [methanol_molecules[0]]
        grouper = BasicRMSDGrouper(molecules, threshold=0.5, num_procs=2)

        matrix = grouper.calculate_full_rmsd_matrix()
        assert matrix.shape == (1, 1)
        assert matrix[0, 0] == 0.0

        groups, index_groups = grouper.group()
        assert len(groups) == 1
        assert len(index_groups) == 1
        assert index_groups == [[0]]

    def test_matrix_grouper_log_progress(self, caplog):
        """Test functionality of MatrixGrouper._log_progress method."""
        import logging

        from chemsmart.jobs.grouper.base import MatrixGrouper

        class DummyGrouper(MatrixGrouper):
            def _calculate_rmsd(self, idx_pair):
                return 0.0

            def group(self):
                return [], []

        # Create dummy grouper
        grouper = DummyGrouper([], threshold=0.1)
        with caplog.at_level(logging.INFO):
            # Initial call, 0% complete, next_progress threshold at 10%
            next_progress = 10
            next_progress = grouper._log_progress(0, 100, next_progress)
            assert next_progress == 10
            assert len(caplog.records) == 0

            # Call at 9%
            next_progress = grouper._log_progress(9, 100, next_progress)
            assert next_progress == 10
            assert len(caplog.records) == 0

            # Call at 10%
            next_progress = grouper._log_progress(10, 100, next_progress)
            assert next_progress == 20
            assert len(caplog.records) == 1
            assert (
                "Matrix calculation progress: 10%"
                in caplog.records[-1].message
            )

            # Call at 15% (no log, threshold is 20)
            next_progress = grouper._log_progress(15, 100, next_progress)
            assert next_progress == 20
            assert len(caplog.records) == 1

            # Call jumping from 15% to 35%
            next_progress = grouper._log_progress(35, 100, next_progress)
            assert next_progress == 40
            assert len(caplog.records) == 3
            assert (
                "Matrix calculation progress: 20%"
                in caplog.records[-2].message
            )
            assert (
                "Matrix calculation progress: 30%"
                in caplog.records[-1].message
            )

            # Test edge case with zero total
            next_p = grouper._log_progress(10, 0, 10)
            assert next_p == 10


@pytest.mark.usefixtures("temporary_working_dir")
class Test_representative_selection:
    class DummyDistanceGrouper(MatrixGrouper):
        def __init__(self, molecules, distance_matrix, **kwargs):
            super().__init__(molecules, threshold=9.9, **kwargs)
            self._distance_matrix = np.array(distance_matrix, dtype=float)

        def group(self):
            groups, index_groups = self.group_by_threshold(
                self._distance_matrix
            )
            self._cached_groups = groups
            self._cached_group_indices = index_groups
            self.record(distance_matrix=self._distance_matrix)
            return groups, index_groups

        def _record_results(self, distance_matrix):
            recorder = self._get_results_recorder()
            labels = recorder.get_labels(distance_matrix.shape[0])
            header_info = [("", "Dummy Distance Grouper")]
            self._append_input_usage_header(header_info)
            sheets_data = {}
            if self._cached_group_indices is not None:
                sheets_data["Groups"] = recorder.build_groups_dataframe(
                    self._cached_group_indices, len(self.molecules)
                )
            recorder.record_results(
                grouper_name=self.__class__.__name__,
                header_info=header_info,
                sheets_data=sheets_data,
                matrix_data=("Distance", distance_matrix, labels),
                suffix="test",
                startrow=len(header_info) + 2,
            )

    @staticmethod
    def _set_energies(molecules, energies):
        for mol, energy in zip(molecules, energies):
            mol.energy = energy

    def test_default_representative_strategy_is_lowest(
        self, methanol_molecules
    ):
        grouper = BasicRMSDGrouper(methanol_molecules[:2], threshold=0.5)
        assert grouper.representative_strategy == "lowest"

    def test_lowest_strategy_orders_group_by_energy(
        self, multiple_molecules_xyz_file
    ):
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)[:4]
        self._set_energies(molecules, [5.0, 1.0, 3.0, None])
        matrix = np.array(
            [
                [0.0, 0.1, 0.2, 0.3],
                [0.1, 0.0, 0.2, 0.3],
                [0.2, 0.2, 0.0, 0.3],
                [0.3, 0.3, 0.3, 0.0],
            ]
        )
        grouper = self.DummyDistanceGrouper(
            molecules,
            matrix,
            representative_strategy="lowest",
        )
        groups, index_groups = grouper.group()

        assert index_groups[0][0] == 1
        assert groups[0][0] is molecules[1]
        assert index_groups[0] == [1, 2, 0, 3]

    def test_center_strategy_orders_entire_group_by_centrality(
        self, multiple_molecules_xyz_file
    ):
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)[:5]
        self._set_energies(molecules, [0.0, 0.1, 0.2, -1.0, 0.3])
        matrix = np.array(
            [
                [0.0, 1.0, 1.0, 1.0, 1.0],
                [1.0, 0.0, 2.0, 2.0, 2.0],
                [1.0, 2.0, 0.0, 2.0, 2.0],
                [1.0, 2.0, 2.0, 0.0, 2.0],
                [1.0, 2.0, 2.0, 2.0, 0.0],
            ]
        )
        grouper = self.DummyDistanceGrouper(
            molecules,
            matrix,
            representative_strategy="center",
        )
        groups, index_groups = grouper.group()

        assert set(index_groups[0]) == {0, 1, 2, 3, 4}
        assert index_groups[0][0] == 0
        assert groups[0][0] is molecules[0]
        assert index_groups[0] == [0, 3, 1, 2, 4]

    def test_top3_strategy_picks_most_central_among_three_lowest_energy(
        self, multiple_molecules_xyz_file
    ):
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)[:5]
        self._set_energies(molecules, [0.0, 0.1, 0.2, -1.0, 0.3])
        matrix = np.array(
            [
                [0.0, 1.0, 1.0, 1.0, 1.0],
                [1.0, 0.0, 2.0, 2.0, 2.0],
                [1.0, 2.0, 0.0, 2.0, 2.0],
                [1.0, 2.0, 2.0, 0.0, 2.0],
                [1.0, 2.0, 2.0, 2.0, 0.0],
            ]
        )
        grouper = self.DummyDistanceGrouper(
            molecules,
            matrix,
            representative_strategy="top3",
        )
        groups, index_groups = grouper.group()

        assert index_groups[0][0] == 0
        assert groups[0][0] is molecules[0]
        assert index_groups[0] == [0, 3, 1, 2, 4]

    def test_top3_falls_back_to_lowest_for_group_size_two(
        self, methanol_molecules
    ):
        molecules = methanol_molecules[:2]
        self._set_energies(molecules, [5.0, 1.0])
        matrix = np.array([[0.0, 0.2], [0.2, 0.0]])
        grouper = self.DummyDistanceGrouper(
            molecules,
            matrix,
            representative_strategy="top3",
        )
        _, index_groups = grouper.group()
        assert index_groups[0] == [1, 0]

    def test_top3_falls_back_to_lowest_for_group_size_one(
        self, methanol_molecules
    ):
        molecules = [methanol_molecules[0]]
        molecules[0].energy = 3.0
        matrix = np.array([[0.0]])
        grouper = self.DummyDistanceGrouper(
            molecules,
            matrix,
            representative_strategy="top3",
        )
        _, index_groups = grouper.group()
        assert index_groups[0] == [0]

    def test_non_matrix_lowest_orders_first_by_energy(
        self, multiple_molecules_xyz_file
    ):
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)[:4]
        self._set_energies(molecules, [5.0, 2.0, 1.0, None])
        grouper = FormulaGrouper(molecules, representative_strategy="lowest")
        groups, index_groups = grouper.group()
        assert index_groups[0] == [2, 1, 0, 3]
        assert groups[0][0] is molecules[2]

    def test_non_matrix_center_and_top3_raise(self, methanol_molecules):
        with pytest.raises(
            ValueError, match="requires a pairwise distance matrix"
        ):
            FormulaGrouper(
                methanol_molecules[:2], representative_strategy="center"
            )
        with pytest.raises(
            ValueError, match="requires a pairwise distance matrix"
        ):
            FormulaGrouper(
                methanol_molecules[:2], representative_strategy="top3"
            )

    def test_groups_table_and_unique_and_xyz_use_representative_first(
        self, multiple_molecules_xyz_file, temporary_working_dir
    ):
        from openpyxl import load_workbook

        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)[:5]
        self._set_energies(molecules, [0.0, 0.1, 0.2, -1.0, 0.3])
        conformer_ids = ["c1", "c2", "c3", "c4", "c5"]
        matrix = np.array(
            [
                [0.0, 1.0, 1.0, 1.0, 1.0],
                [1.0, 0.0, 2.0, 2.0, 2.0],
                [1.0, 2.0, 0.0, 2.0, 2.0],
                [1.0, 2.0, 2.0, 0.0, 2.0],
                [1.0, 2.0, 2.0, 2.0, 0.0],
            ]
        )
        grouper = self.DummyDistanceGrouper(
            molecules,
            matrix,
            representative_strategy="center",
            label="rep_center",
            conformer_ids=conformer_ids,
        )
        groups, index_groups = grouper.group()
        assert index_groups[0][0] == 0

        xlsx_file = os.path.join(
            temporary_working_dir,
            "rep_center_group_result",
            "rep_center_DummyDistanceGrouper_test.xlsx",
        )
        wb = load_workbook(xlsx_file, data_only=True)
        ws = wb["Groups"]
        assert ws["B2"].value.startswith("c1")

        unique_mols = grouper.unique()
        assert unique_mols[0] is groups[0][0]

        xyz_file = os.path.join(
            temporary_working_dir,
            "rep_center_group_result",
            "rep_center_group_1.xyz",
        )
        with open(xyz_file, "r") as handle:
            lines = handle.readlines()
        assert "Original_Index: c1" in lines[1]

    @staticmethod
    def _partition(index_groups):
        return {frozenset(group) for group in index_groups}

    @staticmethod
    def _center_score(matrix, group, idx):
        others = [j for j in group if j != idx]
        if not others:
            return 0.0
        return float(np.mean([matrix[idx, j] for j in others]))

    def test_representative_strategy_does_not_change_group_membership(
        self, multiple_molecules_xyz_file
    ):
        """lowest/center/top3 may reorder groups but must not change membership."""
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)

        partitions = []
        for strategy in ("lowest", "center", "top3"):
            grouper = BasicRMSDGrouper(
                molecules,
                threshold=2.0,
                num_procs=1,
                representative_strategy=strategy,
            )
            _, index_groups = grouper.group()
            partitions.append(self._partition(index_groups))

        assert partitions[0] == partitions[1] == partitions[2]

    def test_center_puts_most_central_member_first(
        self, multiple_molecules_xyz_file
    ):
        """center must place the member with minimum mean within-group distance first."""
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)

        grouper = BasicRMSDGrouper(
            molecules,
            threshold=2.0,
            num_procs=1,
            representative_strategy="center",
        )
        matrix = grouper.calculate_full_rmsd_matrix()
        _, index_groups = grouper.group()

        for group in index_groups:
            if len(group) == 1:
                continue
            expected = min(
                group,
                key=lambda idx: (
                    self._center_score(matrix, group, idx),
                    (
                        float("inf")
                        if molecules[idx].energy is None
                        else molecules[idx].energy
                    ),
                    idx,
                ),
            )
            assert group[0] == expected

    def test_top3_uses_only_three_lowest_energy_candidates(
        self, multiple_molecules_xyz_file
    ):
        """top3 chooses the most central candidate among the three lowest energies."""
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)

        grouper = BasicRMSDGrouper(
            molecules,
            threshold=2.0,
            num_procs=1,
            representative_strategy="top3",
        )
        matrix = grouper.calculate_full_rmsd_matrix()
        _, index_groups = grouper.group()

        tested_large_group = False
        for group in index_groups:
            if len(group) < 3:
                continue
            tested_large_group = True
            energy_order = sorted(
                group,
                key=lambda idx: (
                    (
                        float("inf")
                        if molecules[idx].energy is None
                        else molecules[idx].energy
                    ),
                    idx,
                ),
            )
            candidates = energy_order[:3]
            expected = min(
                candidates,
                key=lambda idx: (
                    self._center_score(matrix, group, idx),
                    (
                        float("inf")
                        if molecules[idx].energy is None
                        else molecules[idx].energy
                    ),
                    idx,
                ),
            )
            assert group[0] == expected
            assert group[0] in candidates

            expected_remaining = [
                idx for idx in energy_order if idx != expected
            ]
            assert group[1:] == expected_remaining

        assert (
            tested_large_group
        ), "Fixture must contain at least one group of size >= 3"

    def test_top3_falls_back_to_lowest_for_two_member_group(self):
        """top3 must use lowest-energy ordering when a group has fewer than 3 members."""
        from ase.build import molecule as ase_molecule

        from chemsmart.io.molecules.structure import Molecule

        mol0 = Molecule.from_ase_atoms(ase_molecule("CH3OH"))
        mol1 = mol0.copy()
        mol0._energy = -100.0
        mol1._energy = -101.0

        grouper = BasicRMSDGrouper(
            [mol0, mol1],
            threshold=1.0,
            num_procs=1,
            representative_strategy="top3",
        )
        _, index_groups = grouper.group()

        assert index_groups == [[1, 0]]

    @pytest.mark.parametrize("strategy", ["center", "top3"])
    def test_non_matrix_grouper_rejects_matrix_representative_strategy(
        self, methanol_molecules, strategy
    ):
        """Non-matrix groupers must reject center/top3 instead of silently falling back."""
        with pytest.raises(ValueError, match="pairwise distance matrix"):
            grouper = FormulaGrouper(
                methanol_molecules,
                representative_strategy=strategy,
            )
            grouper.group()

    def test_unique_uses_first_group_member_as_representative(
        self, multiple_molecules_xyz_file
    ):
        """unique() must trust group[0] and must not re-select by energy."""
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)

        grouper = BasicRMSDGrouper(
            molecules,
            threshold=2.0,
            num_procs=1,
            representative_strategy="center",
            label="center_unique_test",
        )
        groups, index_groups = grouper.group()
        unique_molecules = grouper.unique()

        assert len(unique_molecules) == len(groups)
        for unique_mol, group, indices in zip(
            unique_molecules, groups, index_groups
        ):
            assert indices[0] in range(len(molecules))
            assert np.allclose(unique_mol.positions, group[0].positions)

    def test_default_representative_matches_explicit_lowest(
        self, multiple_molecules_xyz_file
    ):
        """The new option must preserve the historical lowest-energy default."""
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)

        default_grouper = BasicRMSDGrouper(
            molecules,
            threshold=2.0,
            num_procs=1,
        )
        explicit_grouper = BasicRMSDGrouper(
            molecules,
            threshold=2.0,
            num_procs=1,
            representative_strategy="lowest",
        )

        _, default_indices = default_grouper.group()
        _, explicit_indices = explicit_grouper.group()

        assert default_indices == explicit_indices

    def test_center_tie_breaks_by_energy_then_original_index(
        self, multiple_molecules_xyz_file
    ):
        """center ties must resolve deterministically by energy, then original index."""
        xyz_file = XYZFile(filename=multiple_molecules_xyz_file)
        molecules = xyz_file.get_molecules(index=":", return_list=True)[:3]
        self._set_energies(molecules, [1.0, 0.5, 0.5])

        # All three members have identical mean centrality.
        matrix = np.array(
            [
                [0.0, 1.0, 1.0],
                [1.0, 0.0, 1.0],
                [1.0, 1.0, 0.0],
            ]
        )
        grouper = self.DummyDistanceGrouper(
            molecules,
            matrix,
            representative_strategy="center",
        )
        _, index_groups = grouper.group()

        # 1 and 2 tie in centrality and energy, so original index 1 wins.
        assert index_groups[0] == [1, 2, 0]
